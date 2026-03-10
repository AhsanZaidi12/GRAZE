#!/usr/bin/env python3
"""
Enhanced Tackle Segmentation - Segment_tackles.py
============================================================

FIXES:
1. Horizontal player detection: Only reject horizontal DUMMIES, not players
2. Random module conflict: Renamed parameter to avoid shadowing
3. Performance optimization: Balanced validation (fewer frames but smarter)
4. Lower validation thresholds (0.25 vs 0.35)
5. Relaxed matching criteria

srun --time=8:00:00 --nodes=1 --cpus-per-task=4 --mem=64G --gres=gpu:1 --pty bash
python segment_tacklesV2.py     --video_dir /homes/ahsanzaidi/TackleStudy745/videos     --output_dir /homes/ahsanzaidi/Tackle_SAM2_v2     --process_all

python segment_tacklesV2.py     --video_dir /homes/ahsanzaidi/TackleStudy745/videos     --output_dir /homes/ahsanzaidi/Tackle_SAM2_v2     --specific_video 008
"""

import sys
sys.path.insert(0, '/homes/ahsanzaidi/GroundingDINO')

import os
import cv2
import torch
import numpy as np
from pathlib import Path
import argparse
from typing import List, Tuple, Dict, Optional
import json
from datetime import datetime
import random as random_module  # FIXED: Renamed to avoid conflict
from dataclasses import dataclass, field
import traceback
import shutil
import gc
from openpyxl import Workbook, load_workbook

from sam2.build_sam import build_sam2_video_predictor
from groundingdino.util.inference import load_model, load_image, predict


@dataclass
class DetectedObject:
    box: np.ndarray
    score: float
    label: str
    kind: str
    center: Tuple[float, float]
    area: float
    aspect_ratio: float
    frame_idx: int = 0
    proximity_to_dummy: float = float('inf')
    selection_score: float = 0.0

    def to_dict(self):
        return {
            'box': self.box.tolist() if isinstance(self.box, np.ndarray) else self.box,
            'score': float(self.score),
            'label': self.label,
            'kind': self.kind,
            'center': (float(self.center[0]), float(self.center[1])),
            'area': float(self.area),
            'aspect_ratio': float(self.aspect_ratio),
            'frame_idx': int(self.frame_idx),
            'proximity_to_dummy': float(self.proximity_to_dummy),
            'selection_score': float(self.selection_score),
        }


@dataclass
class ValidatedDetection:
    primary_frame: int
    primary_detection: DetectedObject
    validation_detections: List[DetectedObject] = field(default_factory=list)
    consistency_score: float = 0.0
    validated: bool = False
    player_motion_score: float = 0.0
    directional_motion_score: float = 0.0

    def to_dict(self):
        return {
            'primary_frame': int(self.primary_frame),
            'primary_detection': self.primary_detection.to_dict(),
            'validation_detections': [d.to_dict() for d in self.validation_detections],
            'consistency_score': float(self.consistency_score),
            'validated': bool(self.validated),
            'player_motion_score': float(self.player_motion_score),
            'directional_motion_score': float(self.directional_motion_score)
        }


class NumpyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.integer):
            return int(obj)
        if isinstance(obj, np.floating):
            return float(obj)
        if isinstance(obj, np.ndarray):
            return obj.tolist()
        if isinstance(obj, np.bool_):
            return bool(obj)
        return super(NumpyEncoder, self).default(obj)


class ExcelTracker:
    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.wb = None
        self.sheet1 = None
        self.sheet2 = None
        self._load_or_create()

    def _load_or_create(self):
        if os.path.exists(self.excel_path):
            self.wb = load_workbook(self.excel_path)
            if "Correctly Segmented" in self.wb.sheetnames:
                self.sheet1 = self.wb["Correctly Segmented"]
            else:
                self.sheet1 = self.wb.create_sheet("Correctly Segmented")
            if "Need Manual Review" in self.wb.sheetnames:
                self.sheet2 = self.wb["Need Manual Review"]
            else:
                self.sheet2 = self.wb.create_sheet("Need Manual Review")
        else:
            self.wb = Workbook()
            if "Sheet" in self.wb.sheetnames:
                del self.wb["Sheet"]
            self.sheet1 = self.wb.create_sheet("Correctly Segmented")
            self.sheet2 = self.wb.create_sheet("Need Manual Review")
        self._ensure_headers()

    def _ensure_headers(self):
        if self.sheet1.cell(1, 1).value != "Video Filename":
            self.sheet1.insert_rows(1)
            self.sheet1.cell(1, 1).value = "Video Filename"
            self.sheet1.cell(1, 2).value = "Threshold"
            self.sheet1.cell(1, 3).value = "Confidence"
            self.sheet1.cell(1, 4).value = "FPOC Frame"
        if self.sheet2.cell(1, 1).value != "Video Filename":
            self.sheet2.insert_rows(1)
            self.sheet2.cell(1, 1).value = "Video Filename"
            self.sheet2.cell(1, 2).value = "FPOC Frame"

    def _find_row_by_video(self, sheet, video_name: str) -> Optional[int]:
        for row in range(2, sheet.max_row + 1):
            cell_value = sheet.cell(row, 1).value
            if cell_value == video_name:
                return row
        return None

    def add_correctly_segmented(self, video_name: str, threshold: float, confidence: float, fpoc_frame: int):
        row = self._find_row_by_video(self.sheet2, video_name)
        if row:
            self.sheet2.delete_rows(row)
        row = self._find_row_by_video(self.sheet1, video_name)
        if row:
            self.sheet1.cell(row, 2).value = threshold
            self.sheet1.cell(row, 3).value = confidence
            self.sheet1.cell(row, 4).value = fpoc_frame
        else:
            self.sheet1.append([video_name, threshold, confidence, fpoc_frame])

    def add_needs_review(self, video_name: str, fpoc_frame: Optional[int] = None):
        if self._find_row_by_video(self.sheet1, video_name):
            return
        row = self._find_row_by_video(self.sheet2, video_name)
        if row:
            self.sheet2.cell(row, 2).value = fpoc_frame if fpoc_frame else "N/A"
        else:
            self.sheet2.append([video_name, fpoc_frame if fpoc_frame else "N/A"])

    def save(self):
        self.wb.save(self.excel_path)


class ObjectValidator:
    @staticmethod
    def compute_iou(box1: np.ndarray, box2: np.ndarray) -> float:
        x1_1, y1_1, x2_1, y2_1 = [float(v) for v in box1]
        x1_2, y1_2, x2_2, y2_2 = [float(v) for v in box2]
        x1_i = max(x1_1, x1_2)
        y1_i = max(y1_1, y1_2)
        x2_i = min(x2_1, x2_2)
        y2_i = min(y2_1, y2_2)
        if x2_i <= x1_i or y2_i <= y1_i:
            return 0.0
        inter = (x2_i - x1_i) * (y2_i - y1_i)
        a1 = max(0.0, (x2_1 - x1_1)) * max(0.0, (y2_1 - y1_1))
        a2 = max(0.0, (x2_2 - x1_2)) * max(0.0, (y2_2 - y1_2))
        union = a1 + a2 - inter
        return float(inter / union) if union > 1e-9 else 0.0

    @staticmethod
    def center_distance(det1: DetectedObject, det2: DetectedObject) -> float:
        dx = det1.center[0] - det2.center[0]
        dy = det1.center[1] - det2.center[1]
        return float(np.sqrt(dx * dx + dy * dy))

    @staticmethod
    def size_ratio(det1: DetectedObject, det2: DetectedObject) -> float:
        a1 = max(float(det1.area), 1e-9)
        a2 = max(float(det2.area), 1e-9)
        return float(min(a1, a2) / max(a1, a2))

    @staticmethod
    def match_best(primary: DetectedObject, candidates: List[DetectedObject],
                   iou_threshold: float, distance_threshold: float,
                   min_size_ratio: float = 0.45) -> Tuple[Optional[DetectedObject], float]:
        best = None
        best_conf = 0.0
        for cand in candidates:
            if cand.kind != primary.kind:
                continue
            iou = ObjectValidator.compute_iou(primary.box, cand.box)
            dist = ObjectValidator.center_distance(primary, cand)
            sratio = ObjectValidator.size_ratio(primary, cand)
            if sratio < min_size_ratio:
                continue
            if iou < iou_threshold and dist > distance_threshold:
                continue
            iou_score = min(iou / max(iou_threshold, 1e-6), 1.0)
            dist_score = max(0.0, 1.0 - dist / (distance_threshold * 1.25))
            conf = 0.55 * iou_score + 0.30 * dist_score + 0.15 * sratio
            if conf > best_conf:
                best_conf = conf
                best = cand
        return best, float(best_conf)

    @staticmethod
    def validate_across_frames(primary: DetectedObject, matched: List[DetectedObject],
                               min_validations: int = 3, iou_threshold: float = 0.30,
                               distance_threshold: float = 100.0) -> Tuple[bool, float]:
        if not matched:
            return False, 0.0
        confidences = []
        for m in matched:
            iou = ObjectValidator.compute_iou(primary.box, m.box)
            dist = ObjectValidator.center_distance(primary, m)
            sratio = ObjectValidator.size_ratio(primary, m)
            iou_score = min(iou / max(iou_threshold, 1e-6), 1.0)
            dist_score = max(0.0, 1.0 - dist / (distance_threshold * 1.25))
            conf = 0.55 * iou_score + 0.30 * dist_score + 0.15 * sratio
            confidences.append(conf)
        mean_conf = float(np.mean(confidences)) if confidences else 0.0
        validated = len(matched) >= min_validations
        return validated, mean_conf

    @staticmethod
    def compute_player_motion(primary: DetectedObject, matched: List[DetectedObject]) -> float:
        if not matched:
            return 0.0
        movements = []
        for m in matched:
            dx = primary.center[0] - m.center[0]
            dy = primary.center[1] - m.center[1]
            movements.append(np.sqrt(dx * dx + dy * dy))
        avg_movement = float(np.mean(movements))
        return float(min(avg_movement / 200.0, 1.0))

    @staticmethod
    def compute_directional_motion(primary: DetectedObject, matched: List[DetectedObject],
                                   dummy: DetectedObject) -> float:
        if not matched or dummy is None:
            return 0.0
        past_frames = [m for m in matched if m.frame_idx < primary.frame_idx]
        if len(past_frames) < 1:
            return 0.5
        if len(past_frames) == 1:
            start_pos = np.array(past_frames[0].center, dtype=np.float32)
        else:
            past_centers = np.array([m.center for m in past_frames], dtype=np.float32)
            start_pos = np.mean(past_centers, axis=0)
        end_pos = np.array(primary.center, dtype=np.float32)
        dummy_center = np.array(dummy.center, dtype=np.float32)
        motion_vector = end_pos - start_pos
        motion_norm = np.linalg.norm(motion_vector)
        if motion_norm < 1e-6:
            return 0.5
        to_dummy = dummy_center - end_pos
        to_dummy_norm = np.linalg.norm(to_dummy)
        if to_dummy_norm < 1e-6:
            return 0.5
        motion_unit = motion_vector / motion_norm
        to_dummy_unit = to_dummy / to_dummy_norm
        dot_product = float(np.dot(motion_unit, to_dummy_unit))
        score = (dot_product + 1.0) / 2.0
        return float(score)

    @staticmethod
    def make_reference_detection(box: np.ndarray, kind: str, frame_idx: int) -> DetectedObject:
        x1, y1, x2, y2 = [float(v) for v in box]
        center = ((x1 + x2) / 2.0, (y1 + y2) / 2.0)
        area = max((x2 - x1) * (y2 - y1), 1e-6)
        aspect_ratio = (y2 - y1) / (x2 - x1 + 1e-6)
        return DetectedObject(
            box=np.array([x1, y1, x2, y2], dtype=np.float32),
            score=1.0,
            label="reference",
            kind=kind,
            center=center,
            area=float(area),
            aspect_ratio=float(aspect_ratio),
            frame_idx=int(frame_idx)
        )


class ContactValidator:
    @staticmethod
    def compute_mask_overlap(mask1: np.ndarray, mask2: np.ndarray) -> int:
        return int(np.logical_and(mask1, mask2).sum())

    @staticmethod
    def summarize_overlap(overlap_per_frame: Dict[int, int], min_overlap_pixels: int = 1,
                         store_overlap_curve: bool = False, max_overlap: int = 0,
                         fpoc_frame: Optional[int] = None) -> Tuple[Optional[int], Optional[int], List[int], Dict]:
        if not overlap_per_frame:
            metrics = {
                'first_contact_frame': None, 'last_contact_frame': None,
                'num_contact_frames': 0, 'max_overlap_pixels': 0,
                'contact_end_reason': None,
                'overlap_per_frame': {} if store_overlap_curve else None
            }
            return None, None, [], metrics
        frames_sorted = sorted(overlap_per_frame.keys())
        contact_frames = [f for f in frames_sorted if overlap_per_frame[f] >= min_overlap_pixels]
        if not contact_frames:
            metrics = {
                'first_contact_frame': None, 'last_contact_frame': None,
                'num_contact_frames': 0,
                'max_overlap_pixels': max(overlap_per_frame.values()) if overlap_per_frame else 0,
                'contact_end_reason': None,
                'overlap_per_frame': ({int(k): int(v) for k, v in overlap_per_frame.items()} if store_overlap_curve else None)
            }
            return None, None, [], metrics
        first_contact = contact_frames[0]
        max_overlap_val = max(overlap_per_frame.values()) if overlap_per_frame else 0
        last_contact = contact_frames[-1]
        contact_end_reason = "natural_end"
        if fpoc_frame is not None and max_overlap_val > 0:
            overlap_threshold = max_overlap_val * 0.10
            consecutive_low = 0
            required_consecutive = 3
            for frame_idx in range(first_contact, frames_sorted[-1] + 1):
                frames_after_fpoc = frame_idx - fpoc_frame
                current_overlap = overlap_per_frame.get(frame_idx, 0)
                if current_overlap < overlap_threshold:
                    consecutive_low += 1
                else:
                    consecutive_low = 0
                if frames_after_fpoc > 30 and consecutive_low >= required_consecutive:
                    last_contact = frame_idx - required_consecutive
                    contact_end_reason = "dynamic_threshold"
                    break
        contact_frames = [f for f in frames_sorted if f <= last_contact and overlap_per_frame[f] >= min_overlap_pixels]
        metrics = {
            'first_contact_frame': first_contact,
            'last_contact_frame': last_contact,
            'num_contact_frames': len(contact_frames),
            'max_overlap_pixels': int(max_overlap_val),
            'contact_end_reason': contact_end_reason,
            'overlap_per_frame': ({int(k): int(v) for k, v in overlap_per_frame.items()} if store_overlap_curve else None)
        }
        return first_contact, last_contact, contact_frames, metrics


class ImprovedSegmenterAMP:
    GROUNDING_STEP = 5
    MAX_GROUNDING_OFFSET = 50

    PROMPT_SETS = [
        {
            "name": "geared_player",
            "player": "Primary foreground helmeted full-body american football player sprinting, forward-leaning running posture, arms pumping, one leg extended",
            "dummy": "red upright rectangular standing tackle dummy pad, vertical blocking dummy, red standing tackling dummy"
        },
        {
            "name": "non_geared_player",
            "player": "Primary foreground athletic full-body person running, sprinting, forward motion, practicing football tackle",
            "dummy": "red upright rectangular standing tackle dummy pad, vertical blocking dummy"
        },
        {
            "name": "generic_player",
            "player": "person full-body running towards red object, athlete in motion, player approaching target",
            "dummy": "red vertical object, standing red pad, upright red target"
        }
    ]

    def __init__(self, sam2_checkpoint: str, sam2_config: str,
                 grounding_dino_checkpoint: str, grounding_dino_config: str, device: str = "cuda"):
        self.device = device
        self.use_amp = (device == "cuda")
        print("=" * 60)
        print("TACKLE SEGMENTATION - OPTIMIZED VERSION")
        print("=" * 60)
        print("✓ FIXED: Horizontal player detection (only reject horizontal dummies)")
        print("✓ OPTIMIZED: Balanced validation (16 frames, smarter selection)")
        print("✓ Lower validation thresholds (0.25 vs 0.35)")
        print("✓ Relaxed matching for difficult videos")
        print("=" * 60)
        print("\nLoading Grounding DINO model...")
        self.grounding_model = load_model(grounding_dino_config, grounding_dino_checkpoint, device=device)
        self.grounding_model.eval()
        print("✓ Grounding DINO loaded")
        print("\nLoading SAM2 video predictor...")
        self.video_predictor = build_sam2_video_predictor(sam2_config, sam2_checkpoint, device=device)
        print("✓ SAM2 loaded")
        self.validator = ObjectValidator()
        self.contact_validator = ContactValidator()
        if torch.cuda.is_available():
            torch.cuda.empty_cache()

    def detect_objects(self, image_path: str, text_prompt: str,
                      box_threshold: float = 0.30, text_threshold: float = 0.25
                      ) -> Tuple[torch.Tensor, torch.Tensor, List[str], np.ndarray]:
        try:
            image_source, image = load_image(image_path)
            with torch.cuda.amp.autocast(enabled=self.use_amp):
                boxes, logits, phrases = predict(
                    model=self.grounding_model, image=image, caption=text_prompt,
                    box_threshold=box_threshold, text_threshold=text_threshold, device=self.device
                )
            h, w, _ = image_source.shape
            if boxes is None or len(boxes) == 0:
                if torch.cuda.is_available():
                    torch.cuda.empty_cache()
                return torch.empty((0, 4)), torch.empty((0,)), [], image_source
            scale = torch.tensor([w, h, w, h], device=boxes.device, dtype=boxes.dtype)
            boxes_xyxy = boxes * scale
            boxes_xyxy[:, :2] -= boxes_xyxy[:, 2:] / 2
            boxes_xyxy[:, 2:] += boxes_xyxy[:, :2]
            boxes_xyxy[:, 0::2] = torch.clamp(boxes_xyxy[:, 0::2], 0, w - 1)
            boxes_xyxy[:, 1::2] = torch.clamp(boxes_xyxy[:, 1::2], 0, h - 1)
            valid = (boxes_xyxy[:, 2] > boxes_xyxy[:, 0] + 1) & (boxes_xyxy[:, 3] > boxes_xyxy[:, 1] + 1)
            boxes_xyxy = boxes_xyxy[valid]
            logits = logits[valid]
            phrases = [p for p, v in zip(phrases, valid.tolist()) if v]
            if torch.cuda.is_available():
                torch.cuda.empty_cache()
            return boxes_xyxy, logits, phrases, image_source
        except RuntimeError as e:
            if "out of memory" in str(e):
                if torch.cuda.is_available():
                    torch.cuda.empty_cache()
                    gc.collect()
                raise
            raise
        finally:
            if torch.cuda.is_available():
                torch.cuda.empty_cache()

    def parse_detections(self, boxes: torch.Tensor, logits: torch.Tensor, phrases: List[str],
                        image_shape: Tuple[int, int], frame_idx: int) -> List[DetectedObject]:
        detections: List[DetectedObject] = []
        h, w = image_shape
        for box, score, label in zip(boxes, logits, phrases):
            box_np = box.detach().cpu().numpy().astype(np.float32)
            x1, y1, x2, y2 = [float(v) for v in box_np]
            x1 = float(np.clip(x1, 0, w - 1))
            x2 = float(np.clip(x2, 0, w - 1))
            y1 = float(np.clip(y1, 0, h - 1))
            y2 = float(np.clip(y2, 0, h - 1))
            if x2 <= x1 + 1 or y2 <= y1 + 1:
                continue
            center = ((x1 + x2) / 2.0, (y1 + y2) / 2.0)
            area = (x2 - x1) * (y2 - y1)
            aspect_ratio = (y2 - y1) / (x2 - x1 + 1e-6)
            detections.append(DetectedObject(
                box=np.array([x1, y1, x2, y2], dtype=np.float32), score=float(score),
                label=str(label), kind="unknown", center=center, area=float(area),
                aspect_ratio=float(aspect_ratio), frame_idx=int(frame_idx),
            ))
        return detections

    def classify_detections(self, detections: List[DetectedObject], image_shape: Tuple[int, int]
                           ) -> Tuple[List[DetectedObject], List[DetectedObject]]:
        """FIXED: Only reject horizontal DUMMIES, not players!"""
        h, w = image_shape
        player_candidates: List[DetectedObject] = []
        dummy_candidates: List[DetectedObject] = []
        edge_margin_x = w * 0.15
        edge_margin_y = h * 0.10
        min_player_area = (h * w) * 0.01
        
        for det in detections:
            label_l = det.label.lower()
            in_edge_region = (det.center[0] < edge_margin_x or det.center[0] > w - edge_margin_x or det.center[1] < edge_margin_y)
            
            # Classify by label keywords
            is_dummy_label = any(word in label_l for word in ['dummy', 'pad', 'bag', 'tackle', 'training', 'blocking'])
            is_player_label = any(word in label_l for word in ['player', 'person', 'athlete', 'helmet', 'football', 'running', 'runner', 'sprint', 'sprinter'])
            
            # Aspect ratio checks
            is_horizontal = det.aspect_ratio < 0.8  # width > height
            is_vertical = det.aspect_ratio > 2.0    # height >> width
            
            # FIXED: Different logic for dummies vs players
            if is_dummy_label or is_vertical:
                # Dummies should be vertical
                if is_horizontal:
                    continue  # Reject horizontal dummies
                det.kind = "dummy"
                dummy_candidates.append(det)
                continue
            
            # FIXED: Players can be horizontal (tackling!) - don't reject based on aspect ratio
            if is_player_label and not in_edge_region and det.area > min_player_area:
                det.kind = "player"
                player_candidates.append(det)
                # Note: We allow horizontal players - they might be mid-tackle!
        
        return player_candidates, dummy_candidates

    def select_best_dummy(self, dummy_candidates: List[DetectedObject], image_shape: Tuple[int, int]) -> Optional[DetectedObject]:
        if not dummy_candidates:
            return None
        h, w = image_shape
        image_center = (w / 2.0, h / 2.0)
        vertical = [d for d in dummy_candidates if d.aspect_ratio > 2.0]
        if not vertical:
            vertical = [d for d in dummy_candidates if d.aspect_ratio > 1.5]
        if not vertical:
            return None
        for det in vertical:
            dist = float(np.sqrt(((det.center[0] - image_center[0]) / w) ** 2 + ((det.center[1] - image_center[1]) / h) ** 2))
            det.selection_score = float(det.score * 0.4 + (1.0 - dist) * 0.3 + min(det.aspect_ratio / 3.0, 1.0) * 0.3)
        return max(vertical, key=lambda x: x.selection_score)

    def select_players_by_proximity(self, player_candidates: List[DetectedObject], dummy: DetectedObject,
                                   image_shape: Tuple[int, int], top_n: int = 3) -> List[DetectedObject]:
        if not player_candidates:
            return []
        h, w = image_shape
        max_dist = float(np.sqrt(w * w + h * h))
        for p in player_candidates:
            dx = p.center[0] - dummy.center[0]
            dy = p.center[1] - dummy.center[1]
            dist = float(np.sqrt(dx * dx + dy * dy))
            p.proximity_to_dummy = dist
            proximity_score = 1.0 - (dist / max_dist)
            size_score = min(p.area / (h * w * 0.2), 1.0)
            confidence_score = p.score
            p.selection_score = float(proximity_score * 0.60 + confidence_score * 0.25 + size_score * 0.15)
        return sorted(player_candidates, key=lambda x: x.selection_score, reverse=True)[:top_n]

    def _grounding_search_offsets(self) -> List[int]:
        offsets = [0]
        for d in range(self.GROUNDING_STEP, self.MAX_GROUNDING_OFFSET + 1, self.GROUNDING_STEP):
            offsets.append(d)
            offsets.append(-d)
        return offsets

    def find_grounding_frame_multiposition(
        self,
        frame_paths: List[str],
        player_prompt: str,
        dummy_prompt: str,
        box_threshold: float = 0.30,
        text_threshold: float = 0.25,
        position_fraction: float = 0.5
    ) -> Tuple[int, List[ValidatedDetection], Optional[ValidatedDetection]]:
        """
        OPTIMIZED: Fewer but smarter validation frames (16 instead of 24)
        """
        num_frames = len(frame_paths)
        middle = int(num_frames * position_fraction)
        combined_prompt = f"{player_prompt} . {dummy_prompt}"
        
        primary_frame: Optional[int] = None
        best_dummy: Optional[DetectedObject] = None
        top_players: List[DetectedObject] = []

        for off in self._grounding_search_offsets():
            idx = middle + off
            if idx < 0 or idx >= num_frames:
                continue
            try:
                boxes, logits, phrases, image = self.detect_objects(frame_paths[idx], combined_prompt, box_threshold, text_threshold)
                if len(boxes) == 0:
                    continue
                shape = (image.shape[0], image.shape[1])
                dets = self.parse_detections(boxes, logits, phrases, shape, idx)
                players, dummies = self.classify_detections(dets, shape)
                dummy = self.select_best_dummy(dummies, shape)
                if dummy is None or not players:
                    continue
                top_players = self.select_players_by_proximity(players, dummy, shape, top_n=3)
                if not top_players:
                    continue
                primary_frame = idx
                best_dummy = dummy
                break
            except Exception:
                if torch.cuda.is_available():
                    torch.cuda.empty_cache()
                continue

        if primary_frame is None or best_dummy is None or not top_players:
            if torch.cuda.is_available():
                torch.cuda.empty_cache()
            return -1, [], None

        # OPTIMIZED: 16 validation frames (was 24) - balanced performance
        validation_offsets = [3, -3, 5, -5, 8, -8, 10, -10, 15, -15, 20, -20, 25, -25, 30, -30 , 35, -35]
        validation_frames = [primary_frame + o for o in validation_offsets if 0 <= primary_frame + o < num_frames]

        # Use LOWER thresholds for validation (easier to find matches)
        val_box_thresh = max(0.20, box_threshold - 0.10)
        val_text_thresh = max(0.20, text_threshold - 0.05)

        cached_players: Dict[int, List[DetectedObject]] = {}
        cached_dummies: Dict[int, List[DetectedObject]] = {}

        for vf in validation_frames:
            try:
                boxes, logits, phrases, image = self.detect_objects(
                    frame_paths[vf], combined_prompt, val_box_thresh, val_text_thresh
                )
                if len(boxes) == 0:
                    cached_players[vf] = []
                    cached_dummies[vf] = []
                    continue
                shape = (image.shape[0], image.shape[1])
                dets = self.parse_detections(boxes, logits, phrases, shape, vf)
                p_cands, d_cands = self.classify_detections(dets, shape)
                cached_players[vf] = p_cands
                cached_dummies[vf] = d_cands
            except Exception:
                cached_players[vf] = []
                cached_dummies[vf] = []
                if torch.cuda.is_available():
                    torch.cuda.empty_cache()

        # Dummy validation - RELAXED thresholds
        dummy_matches: List[DetectedObject] = []
        for vf in validation_frames:
            match, _ = self.validator.match_best(
                best_dummy,
                cached_dummies.get(vf, []),
                iou_threshold=0.20,
                distance_threshold=200.0,
                min_size_ratio=0.35
            )
            if match is not None:
                dummy_matches.append(match)

        dummy_validated, dummy_consistency = self.validator.validate_across_frames(
            best_dummy, dummy_matches, min_validations=2, iou_threshold=0.20, distance_threshold=200.0
        )
        validated_dummy = ValidatedDetection(primary_frame, best_dummy, dummy_matches, dummy_consistency, dummy_validated)

        # Player validation - RELAXED thresholds
        validated_players: List[ValidatedDetection] = []
        for p0 in top_players:
            matches: List[DetectedObject] = []
            for vf in validation_frames:
                match, _ = self.validator.match_best(
                    p0,
                    cached_players.get(vf, []),
                    iou_threshold=0.20,
                    distance_threshold=200.0,
                    min_size_ratio=0.35
                )
                if match is not None:
                    matches.append(match)

            p_valid, p_cons = self.validator.validate_across_frames(
                p0, matches, min_validations=3, iou_threshold=0.20, distance_threshold=200.0
            )
            p_motion = self.validator.compute_player_motion(p0, matches)
            p_directional = self.validator.compute_directional_motion(p0, matches, best_dummy)

            vp = ValidatedDetection(primary_frame, p0, matches, p_cons, p_valid, p_motion, p_directional)
            validated_players.append(vp)

        if torch.cuda.is_available():
            torch.cuda.empty_cache()

        return primary_frame, validated_players, validated_dummy

    def find_grounding_frame_progressive(
        self,
        frame_paths: List[str]
    ) -> Tuple[int, List[ValidatedDetection], Optional[ValidatedDetection], str, float]:
        """Proper rejection logic - reject if insufficient validation"""
        print("\n" + "=" * 60)
        print("PHASE 1: Progressive Multi-Position Grounding")
        print("=" * 60)

        positions = [0.2, 0.4, 0.5, 0.6, 0.8, 0.9]
        thresholds = [0.35, 0.30, 0.28]

        for prompt_set in self.PROMPT_SETS:
            print(f"\n🎯 Trying prompt set: {prompt_set['name']}")
            player_prompt = prompt_set['player']
            dummy_prompt = prompt_set['dummy']

            for threshold in thresholds:
                print(f"\n📊 Threshold: {threshold:.2f}")

                for position in positions:
                    print(f"\n  Trying position {position:.2f} (frame {int(len(frame_paths) * position)})...")
                    frame, players, dummy = self.find_grounding_frame_multiposition(
                        frame_paths, player_prompt, dummy_prompt, threshold, 0.25, position
                    )

                    if frame == -1 or not players or dummy is None:
                        continue

                    if len(dummy.validation_detections) < 2:
                        print(f"    ⚠️ Dummy validation failed (validated=False, consistency={dummy.consistency_score:.3f})")
                        continue

                    players_kept: List[ValidatedDetection] = []
                    for p in players:
                        num_matches = len(p.validation_detections)
                        
                        if num_matches < 3:
                            continue
                        if p.player_motion_score < 0.08:
                            continue
                        if p.directional_motion_score < 0.30:
                            continue
                        if p.consistency_score < 0.20:
                            continue
                        
                        players_kept.append(p)

                    if players_kept:
                        players_kept.sort(
                            key=lambda p: (p.directional_motion_score, p.player_motion_score, p.consistency_score),
                            reverse=True
                        )
                        print(f"    ✓ Found both objects at frame {frame}")
                        print(f"\n✓ SUCCESS with {prompt_set['name']}, threshold {threshold:.2f}, position {position:.2f}")
                        print(f"   Player motion: {players_kept[0].player_motion_score:.3f}")
                        print(f"   Directional motion: {players_kept[0].directional_motion_score:.3f}")
                        print(f"   Consistency: {players_kept[0].consistency_score:.3f}")
                        return frame, players_kept, dummy, prompt_set['name'], threshold

                    # Diagnostics
                    print(f"    ✓ Found both objects at frame {frame}")
                    for idx, p in enumerate(players):
                        reasons = []
                        if len(p.validation_detections) < 3:
                            reasons.append(f"low consistency ({p.consistency_score:.3f})")
                        if p.player_motion_score < 0.08:
                            reasons.append(f"low motion ({p.player_motion_score:.3f})")
                        if p.directional_motion_score < 0.30:
                            reasons.append(f"low directional ({p.directional_motion_score:.3f})")
                        if reasons:
                            print(f"    ⚠️ Rejected: {'; '.join(reasons)}")

        print("\n❌ All grounding attempts failed!")
        return -1, [], None, "none", 0.35

    def enhanced_backward_refinement(self, frame_paths, initial_frame, player_prompt, dummy_prompt,
                                    box_threshold, text_threshold=0.25,
                                    reference_player_box=None, reference_dummy_box=None) -> int:
        print("\n🔍 Enhanced backward refinement...")
        combined_prompt = f"{player_prompt} . {dummy_prompt}"
        true_first = initial_frame
        misses = 0
        max_backward = 20
        stop = max(0, initial_frame - max_backward)

        for idx in range(initial_frame - 1, stop - 1, -1):
            try:
                boxes, logits, phrases, image = self.detect_objects(frame_paths[idx], combined_prompt, box_threshold, text_threshold)
                if len(boxes) == 0:
                    misses += 1
                    if misses > 1:
                        break
                    continue
                shape = (image.shape[0], image.shape[1])
                dets = self.parse_detections(boxes, logits, phrases, shape, idx)
                players, dummies = self.classify_detections(dets, shape)

                if reference_player_box is not None and reference_dummy_box is not None:
                    ref_player = self.validator.make_reference_detection(reference_player_box, "player", idx)
                    ref_dummy = self.validator.make_reference_detection(reference_dummy_box, "dummy", idx)
                    player_match, _ = self.validator.match_best(ref_player, players, 0.25, 150.0)
                    dummy_match, _ = self.validator.match_best(ref_dummy, dummies, 0.25, 150.0)
                    if player_match and dummy_match:
                        true_first = idx
                        misses = 0
                    else:
                        misses += 1
                        if misses > 1:
                            break
                else:
                    if players and dummies:
                        true_first = idx
                        misses = 0
                    else:
                        misses += 1
                        if misses > 1:
                            break
            except Exception:
                misses += 1
                if misses > 1:
                    break

        recovered = initial_frame - true_first
        if recovered > 0:
            print(f"   Phase 1: Refined to frame {true_first} (recovered {recovered} frames)")

        print(f"   Phase 2: Gap detection...")
        gap_offsets = [5, 10, 20, 50]
        earliest_gap_frame = true_first

        for gap in gap_offsets:
            test_frame = max(0, true_first - gap)
            if test_frame < 0:
                break
            try:
                boxes, logits, phrases, image = self.detect_objects(frame_paths[test_frame], combined_prompt, box_threshold, text_threshold)
                if len(boxes) == 0:
                    continue
                shape = (image.shape[0], image.shape[1])
                dets = self.parse_detections(boxes, logits, phrases, shape, test_frame)
                players, dummies = self.classify_detections(dets, shape)

                objects_found = False
                if reference_player_box is not None and reference_dummy_box is not None:
                    ref_player = self.validator.make_reference_detection(reference_player_box, "player", test_frame)
                    ref_dummy = self.validator.make_reference_detection(reference_dummy_box, "dummy", test_frame)
                    player_match, _ = self.validator.match_best(ref_player, players, 0.25, 150.0)
                    dummy_match, _ = self.validator.match_best(ref_dummy, dummies, 0.25, 150.0)
                    objects_found = (player_match is not None and dummy_match is not None)
                else:
                    objects_found = (players and dummies)

                if objects_found:
                    print(f"      ✓ Found both at frame {test_frame} (gap={gap})")
                    earliest_gap_frame = test_frame
                    left = test_frame
                    right = true_first
                    while right - left > 1:
                        mid = (left + right) // 2
                        boxes, logits, phrases, image = self.detect_objects(frame_paths[mid], combined_prompt, box_threshold, text_threshold)
                        if len(boxes) > 0:
                            shape = (image.shape[0], image.shape[1])
                            dets = self.parse_detections(boxes, logits, phrases, shape, mid)
                            players, dummies = self.classify_detections(dets, shape)

                            mid_objects_found = False
                            if reference_player_box is not None and reference_dummy_box is not None:
                                ref_player = self.validator.make_reference_detection(reference_player_box, "player", mid)
                                ref_dummy = self.validator.make_reference_detection(reference_dummy_box, "dummy", mid)
                                player_match, _ = self.validator.match_best(ref_player, players, 0.25, 150.0)
                                dummy_match, _ = self.validator.match_best(ref_dummy, dummies, 0.25, 150.0)
                                mid_objects_found = (player_match is not None and dummy_match is not None)
                            else:
                                mid_objects_found = (players and dummies)

                            if mid_objects_found:
                                earliest_gap_frame = mid
                                right = mid
                            else:
                                left = mid
                        else:
                            left = mid
                    true_first = earliest_gap_frame
                    break
            except Exception:
                continue

        total_recovered = initial_frame - true_first
        print(f"   ✓ Final refined frame: {true_first} (recovered {total_recovered} frames)")

        if torch.cuda.is_available():
            torch.cuda.empty_cache()
        return true_first

    def segment_video(self, video_path, output_dir, min_overlap_pixels=1,
                     store_overlap_curve=False, save_visualization=True,
                     save_masks=True, extract_fpoc=True) -> dict:
        video_name = Path(video_path).stem
        print(f"\n{'='*60}")
        print(f"Processing: {video_name}")
        print("=" * 60)

        video_output_dir = os.path.join(output_dir, video_name)
        os.makedirs(video_output_dir, exist_ok=True)
        frames_dir = os.path.join(video_output_dir, "frames")
        os.makedirs(frames_dir, exist_ok=True)

        print("\nExtracting frames...")
        frame_paths = self._extract_frames(video_path, frames_dir)
        if not frame_paths:
            return {"success": False, "error": "No frames extracted", "video_name": video_name}
        print(f"Extracted {len(frame_paths)} frames")

        try:
            grounding_frame, validated_players, validated_dummy, prompt_name, threshold = self.find_grounding_frame_progressive(frame_paths)

            if grounding_frame == -1 or not validated_players or validated_dummy is None:
                print(f"   ✓ Added '{video_name}' to 'Need Manual Review'")
                return {
                    "success": False,
                    "error": "No objects detected after all attempts",
                    "video_name": video_name,
                    "needs_manual_review": True
                }

            prompt_set = next((p for p in self.PROMPT_SETS if p['name'] == prompt_name), self.PROMPT_SETS[0])
            player_prompt = prompt_set['player']
            dummy_prompt = prompt_set['dummy']

            top_player_box = validated_players[0].primary_detection.box
            dummy_box = validated_dummy.primary_detection.box

            refined_first_frame = self.enhanced_backward_refinement(
                frame_paths, grounding_frame, player_prompt, dummy_prompt, threshold, 0.25,
                reference_player_box=top_player_box,
                reference_dummy_box=dummy_box
            )

            players_sorted = validated_players
            actual_grounding_frame = grounding_frame

            if refined_first_frame < grounding_frame:
                try:
                    boxes, logits, phrases, image = self.detect_objects(
                        frame_paths[refined_first_frame], f"{player_prompt} . {dummy_prompt}", threshold, 0.25)
                    if len(boxes) > 0:
                        shape = (image.shape[0], image.shape[1])
                        dets = self.parse_detections(boxes, logits, phrases, shape, refined_first_frame)
                        p_ref, d_ref = self.classify_detections(dets, shape)
                        if p_ref:
                            orig = players_sorted[0].primary_detection
                            best_p = min(p_ref, key=lambda p: np.hypot(p.center[0] - orig.center[0], p.center[1] - orig.center[1]))
                            top_player_box = best_p.box
                        if d_ref:
                            orig_d = validated_dummy.primary_detection
                            best_d = min(d_ref, key=lambda d: np.hypot(d.center[0] - orig_d.center[0], d.center[1] - orig_d.center[1]))
                            dummy_box = best_d.box
                        actual_grounding_frame = refined_first_frame
                except Exception:
                    pass

            best_result = None
            best_idx = 0

            for idx, vp in enumerate(players_sorted):
                seed_frame = actual_grounding_frame if idx == 0 else grounding_frame
                player_seed_box = top_player_box if idx == 0 else vp.primary_detection.box
                dummy_seed_box = dummy_box if idx == 0 else validated_dummy.primary_detection.box

                stats = self._propagate_and_analyze(
                    frames_dir, len(frame_paths), seed_frame,
                    player_seed_box, dummy_seed_box,
                    False, None,
                    min_overlap_pixels, store_overlap_curve
                )

                overall_confidence = (
                    vp.consistency_score * 0.3 +
                    vp.player_motion_score * 0.3 +
                    vp.directional_motion_score * 0.4
                )

                result = {
                    "video_name": video_name,
                    "num_frames": len(frame_paths),
                    "grounding_frame": grounding_frame,
                    "refined_first_frame": refined_first_frame,
                    "actual_grounding_frame": seed_frame,
                    "player_candidate_index": idx,
                    "prompt_set_used": prompt_name,
                    "threshold_used": float(threshold),
                    "player_validated": vp.to_dict(),
                    "dummy_validated": validated_dummy.to_dict(),
                    "overall_confidence": float(overall_confidence),
                    "first_frame_both_objects": stats.get("first_frame_both_objects"),
                    "last_frame_both_objects": stats.get("last_frame_both_objects"),
                    "first_contact_frame": stats.get("first_contact_frame"),
                    "last_contact_frame": stats.get("last_contact_frame"),
                    "contact_frames": stats.get("contact_frames", []),
                    "contact_metrics": stats.get("contact_metrics", {}),
                    "success": True,
                    "has_contact": stats.get("first_contact_frame") is not None,
                    "needs_manual_review": False,
                    "amp_mode": True
                }

                if not result["has_contact"]:
                    result["needs_manual_review"] = True
                elif overall_confidence < 0.35:
                    result["needs_manual_review"] = True
                elif vp.directional_motion_score < 0.30:
                    result["needs_manual_review"] = True

                if result["has_contact"] and not result["needs_manual_review"]:
                    best_result = result
                    best_idx = idx
                    break
                if best_result is None:
                    best_result = result
                    best_idx = idx

                if torch.cuda.is_available():
                    torch.cuda.empty_cache()
                gc.collect()

            if best_result is None:
                print(f"   ✓ Added '{video_name}' to 'Need Manual Review'")
                return {
                    "success": False,
                    "error": "All candidates failed",
                    "video_name": video_name,
                    "needs_manual_review": True
                }

            if save_masks or save_visualization:
                best_player = players_sorted[best_idx]
                seed_frame = actual_grounding_frame if best_idx == 0 else grounding_frame
                player_seed_box = top_player_box if best_idx == 0 else best_player.primary_detection.box
                dummy_seed_box = dummy_box if best_idx == 0 else validated_dummy.primary_detection.box

                masks_dir = os.path.join(video_output_dir, "masks")
                vis_dir = os.path.join(video_output_dir, "visualization")

                if os.path.exists(masks_dir):
                    shutil.rmtree(masks_dir, ignore_errors=True)
                os.makedirs(masks_dir, exist_ok=True)

                if save_visualization:
                    if os.path.exists(vis_dir):
                        shutil.rmtree(vis_dir, ignore_errors=True)
                    os.makedirs(vis_dir, exist_ok=True)

                _ = self._propagate_and_analyze(
                    frames_dir, len(frame_paths), seed_frame,
                    player_seed_box, dummy_seed_box,
                    True, masks_dir,
                    min_overlap_pixels, False
                )

                if save_masks:
                    best_result["masks_dir"] = masks_dir

                if save_visualization:
                    self._save_visualization_from_masks(
                        frame_paths, masks_dir, vis_dir,
                        best_result.get("actual_grounding_frame", grounding_frame),
                        best_result.get("first_frame_both_objects"),
                        best_result.get("last_frame_both_objects"),
                        best_result.get("first_contact_frame"),
                        best_result.get("last_contact_frame"),
                        best_result.get("contact_frames", []),
                        ["Player", "Dummy"]
                    )
                    best_result["visualization_dir"] = vis_dir

            if extract_fpoc and best_result.get("first_contact_frame") is not None:
                fpoc_dir = os.path.join(output_dir, "fpoc_frames")
                os.makedirs(fpoc_dir, exist_ok=True)
                fpoc_frame_idx = best_result["first_contact_frame"]
                fpoc_source = os.path.join(frames_dir, f"{fpoc_frame_idx:05d}.jpg")
                fpoc_dest = os.path.join(fpoc_dir, f"{video_name}_fpoc_frame{fpoc_frame_idx:05d}.jpg")
                if os.path.exists(fpoc_source):
                    shutil.copy2(fpoc_source, fpoc_dest)
                    best_result["fpoc_frame_path"] = fpoc_dest
                    print(f"✓ FPOC frame extracted: {fpoc_dest}")

            metadata_path = os.path.join(video_output_dir, "metadata.json")
            with open(metadata_path, 'w') as f:
                json.dump(best_result, f, indent=2, cls=NumpyEncoder)

            print(f"\n✓ SUCCESS")
            print(f"Overall Confidence: {best_result['overall_confidence']:.3f}")
            print(f"Directional Motion: {validated_players[best_idx].directional_motion_score:.3f}")
            print(f"Consistency: {validated_players[best_idx].consistency_score:.3f}")
            print(f"FPOC: Frame {best_result.get('first_contact_frame', 'N/A')}")
            print(f"Last Contact: Frame {best_result.get('last_contact_frame', 'N/A')} ({best_result.get('contact_metrics', {}).get('contact_end_reason', 'N/A')})")
            print(f"Needs Manual Review: {best_result.get('needs_manual_review', False)}")
            if best_result.get('needs_manual_review'):
                print(f"   ✓ Added '{video_name}' to 'Need Manual Review'")
            else:
                print(f"   ✓ Added '{video_name}' to 'Correctly Segmented'")

            return best_result

        except Exception as e:
            traceback.print_exc()
            return {
                "success": False,
                "error": str(e),
                "video_name": video_name,
                "needs_manual_review": True
            }
        finally:
            if torch.cuda.is_available():
                torch.cuda.empty_cache()
                gc.collect()

    def _propagate_and_analyze(self, frames_dir, num_frames, seed_frame,
                               player_box, dummy_box,
                               save_masks, masks_dir,
                               min_overlap_pixels, store_overlap_curve) -> dict:
        player_box = player_box.astype(np.float32)
        dummy_box = dummy_box.astype(np.float32)
        inference_state = self.video_predictor.init_state(video_path=frames_dir, async_loading_frames=True)
        self.video_predictor.add_new_points_or_box(
            inference_state=inference_state,
            frame_idx=int(seed_frame),
            obj_id=0,
            box=player_box
        )
        self.video_predictor.add_new_points_or_box(
            inference_state=inference_state,
            frame_idx=int(seed_frame),
            obj_id=1,
            box=dummy_box
        )

        if save_masks and masks_dir:
            os.makedirs(os.path.join(masks_dir, "object_0"), exist_ok=True)
            os.makedirs(os.path.join(masks_dir, "object_1"), exist_ok=True)

        first_both = None
        last_both = None
        overlap_per_frame: Dict[int, int] = {}
        max_overlap = 0

        for out_frame_idx, out_obj_ids, out_mask_logits in self.video_predictor.propagate_in_video(inference_state):
            mask_map: Dict[int, np.ndarray] = {}
            for i, out_obj_id in enumerate(out_obj_ids):
                oid = int(out_obj_id)
                if oid not in (0, 1):
                    continue
                mask = (out_mask_logits[i] > 0.0).detach().cpu().numpy().squeeze().astype(bool)
                mask_map[oid] = mask
                if save_masks and masks_dir:
                    cv2.imwrite(
                        os.path.join(masks_dir, f"object_{oid}", f"{int(out_frame_idx):05d}.png"),
                        (mask.astype(np.uint8) * 255)
                    )

            has_both = (0 in mask_map) and (1 in mask_map)
            if has_both:
                if first_both is None:
                    first_both = int(out_frame_idx)
                last_both = int(out_frame_idx)
                ov = self.contact_validator.compute_mask_overlap(mask_map[0], mask_map[1])
                overlap_per_frame[int(out_frame_idx)] = int(ov)
                max_overlap = max(max_overlap, ov)

            if int(out_frame_idx) % 50 == 0 and torch.cuda.is_available():
                torch.cuda.empty_cache()

        first_contact_temp, _, _, _ = self.contact_validator.summarize_overlap(
            overlap_per_frame, min_overlap_pixels, False, max_overlap, None)
        first_contact, last_contact, contact_frames, contact_metrics = self.contact_validator.summarize_overlap(
            overlap_per_frame, min_overlap_pixels, store_overlap_curve, max_overlap, first_contact_temp)

        return {
            "first_frame_both_objects": first_both,
            "last_frame_both_objects": last_both,
            "first_contact_frame": first_contact,
            "last_contact_frame": last_contact,
            "contact_frames": contact_frames,
            "contact_metrics": contact_metrics
        }

    def _extract_frames(self, video_path, output_dir) -> List[str]:
        out = Path(output_dir)
        existing = sorted(out.glob("*.jpg"))
        if len(existing) > 5:
            return [str(p) for p in existing]
        for p in existing:
            try:
                p.unlink()
            except Exception:
                pass
        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened():
            return []
        paths: List[str] = []
        idx = 0
        while True:
            ret, frame = cap.read()
            if not ret:
                break
            fp = os.path.join(output_dir, f"{idx:05d}.jpg")
            if cv2.imwrite(fp, frame):
                paths.append(fp)
                idx += 1
            else:
                break
        cap.release()
        return paths

    def _save_visualization_from_masks(self, frame_paths, masks_dir, output_dir, grounding_frame_idx,
                                      first_frame_both, last_frame_both, first_contact_frame,
                                      last_contact_frame, contact_frames, object_labels):
        colors = [(0, 255, 255), (255, 0, 0)]
        contact_set = set(contact_frames)
        obj0_dir = Path(masks_dir) / "object_0"
        obj1_dir = Path(masks_dir) / "object_1"
        if not obj0_dir.exists() or not obj1_dir.exists():
            return
        obj0_frames = {int(p.stem) for p in obj0_dir.glob("*.png")}
        obj1_frames = {int(p.stem) for p in obj1_dir.glob("*.png")}
        frames_with_both = sorted(obj0_frames.intersection(obj1_frames))

        for frame_idx in frames_with_both:
            if frame_idx < 0 or frame_idx >= len(frame_paths):
                continue
            frame = cv2.imread(frame_paths[frame_idx])
            if frame is None:
                continue
            m0 = cv2.imread(str(obj0_dir / f"{frame_idx:05d}.png"), cv2.IMREAD_GRAYSCALE)
            m1 = cv2.imread(str(obj1_dir / f"{frame_idx:05d}.png"), cv2.IMREAD_GRAYSCALE)
            if m0 is None or m1 is None:
                continue

            masks = {0: (m0 > 0).astype(np.uint8), 1: (m1 > 0).astype(np.uint8)}
            overlay = frame.copy()
            is_first_contact = (first_contact_frame is not None and frame_idx == first_contact_frame)
            is_last_contact = (last_contact_frame is not None and frame_idx == last_contact_frame)
            is_contact = (frame_idx in contact_set)

            for obj_id, mask_binary in masks.items():
                color = colors[obj_id % len(colors)]
                if is_contact:
                    color = tuple(min(255, int(c * 1.3)) for c in color)
                overlay[mask_binary > 0] = color
                contours, _ = cv2.findContours(mask_binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                if contours:
                    c = max(contours, key=cv2.contourArea)
                    x, y, w, h = cv2.boundingRect(c)
                    label = object_labels[obj_id] if obj_id < len(object_labels) else f"Object {obj_id}"
                    thickness = 4 if (is_first_contact or is_last_contact) else (3 if is_contact else 2)
                    cv2.rectangle(overlay, (x, y), (x + w, y + h), color, thickness)
                    label_text = label
                    if is_first_contact:
                        label_text += " [FPOC!]"
                    elif is_last_contact:
                        label_text += " [LAST CONTACT]"
                    elif is_contact:
                        label_text += " [CONTACT]"
                    label_size, _ = cv2.getTextSize(label_text, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)
                    cv2.rectangle(overlay, (x, y - label_size[1] - 10), (x + label_size[0], y), color, -1)
                    cv2.putText(overlay, label_text, (x, y - 5), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

            vis = cv2.addWeighted(frame, 0.6, overlay, 0.4, 0)
            if is_first_contact:
                cv2.putText(vis, "*** FIRST POINT OF CONTACT (FPOC) ***", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 255, 255), 3)
            elif is_last_contact:
                cv2.putText(vis, "*** LAST POINT OF CONTACT ***", (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 0, 255), 3)
            cv2.imwrite(os.path.join(output_dir, f"{frame_idx:05d}.jpg"), vis)


def process_videos(video_dir, output_dir, num_test=None, random_selection=False,
                  specific_video=None, process_all=False, batch_mode=False, batch_id=0, batch_size=100,
                  sam2_checkpoint="weights/sam2.1_hiera_large.pt",
                  grounding_checkpoint="weights/groundingdino_swint_ogc.pth",
                  grounding_config="weights/GroundingDINO_SwinT_OGC.py",
                  min_overlap_pixels=1, store_overlap_curve=False,
                  save_visualization=True, save_masks=True, extract_fpoc=True):
    """FIXED: random_selection parameter to avoid module conflict"""

    segmenter = ImprovedSegmenterAMP(
        sam2_checkpoint, "configs/sam2.1/sam2.1_hiera_l.yaml",
        grounding_checkpoint, grounding_config,
        "cuda" if torch.cuda.is_available() else "cpu"
    )

    video_files = sorted([f for f in os.listdir(video_dir)
                         if f.endswith(('.mp4', '.avi', '.mov', '.MP4', '.AVI', '.MOV'))])

    excel_path = os.path.join(output_dir, "video_tracking.xlsx")
    tracker = ExcelTracker(excel_path)

    if specific_video is not None:
        selected = [f for f in video_files if specific_video in f]
    elif batch_mode:
        start_idx = batch_id * batch_size
        end_idx = min(start_idx + batch_size, len(video_files))
        selected = video_files[start_idx:end_idx]
    elif process_all:
        selected = video_files
    elif random_selection and num_test:
        selected = random_module.sample(video_files, min(num_test, len(video_files)))  # FIXED: Use random_module
    elif num_test:
        selected = video_files[:num_test]
    else:
        selected = video_files[:2]

    results = []
    for i, vf in enumerate(selected):
        print(f"\n{'='*60}\nVIDEO {i+1}/{len(selected)}: {vf}\n{'='*60}")
        vp = os.path.join(video_dir, vf)
        res = segmenter.segment_video(
            vp, output_dir,
            min_overlap_pixels=min_overlap_pixels,
            store_overlap_curve=store_overlap_curve,
            save_visualization=save_visualization,
            save_masks=save_masks,
            extract_fpoc=extract_fpoc
        )
        results.append(res)
        video_name = res.get("video_name", vf)
        if res.get("success") and not res.get("needs_manual_review"):
            tracker.add_correctly_segmented(
                video_name=video_name,
                threshold=res.get("threshold_used", 0.0),
                confidence=res.get("overall_confidence", 0.0),
                fpoc_frame=res.get("first_contact_frame", 0)
            )
        else:
            tracker.add_needs_review(
                video_name=video_name,
                fpoc_frame=res.get("first_contact_frame")
            )
        tracker.save()
        print(f"\n📊 Excel file saved: {excel_path}")
        if torch.cuda.is_available():
            torch.cuda.empty_cache()
        gc.collect()

    summary_path = os.path.join(output_dir, f"summary_batch{batch_id if batch_mode else 0}.json")
    with open(summary_path, 'w') as f:
        json.dump({
            "timestamp": datetime.now().isoformat(),
            "num_videos_tested": len(results),
            "success_rate": (sum(1 for r in results if r.get('success')) / len(results)) if results else 0,
            "contact_detection_rate": (sum(1 for r in results if r.get('has_contact')) / len(results)) if results else 0,
            "correctly_segmented_rate": (sum(1 for r in results if r.get('success') and not r.get('needs_manual_review')) / len(results)) if results else 0,
            "results": results
        }, f, indent=2, cls=NumpyEncoder)


def main():
    parser = argparse.ArgumentParser(description="Enhanced tackle segmentation - OPTIMIZED")
    parser.add_argument("--video_dir", type=str, required=True)
    parser.add_argument("--output_dir", type=str, required=True)
    parser.add_argument("--num_test", type=int, default=None)
    parser.add_argument("--random", action="store_true", dest="random_selection")  # FIXED: dest parameter
    parser.add_argument("--specific_video", type=str, default=None)
    parser.add_argument("--process_all", action="store_true")
    parser.add_argument("--batch_mode", action="store_true")
    parser.add_argument("--batch_id", type=int, default=0)
    parser.add_argument("--batch_size", type=int, default=100)
    parser.add_argument("--sam2_checkpoint", type=str, default="weights/sam2.1_hiera_large.pt")
    parser.add_argument("--grounding_checkpoint", type=str, default="weights/groundingdino_swint_ogc.pth")
    parser.add_argument("--grounding_config", type=str, default="weights/GroundingDINO_SwinT_OGC.py")
    parser.add_argument("--min_overlap_pixels", type=int, default=1)
    parser.add_argument("--store_overlap_curve", action="store_true")
    parser.add_argument("--save_visualization", type=lambda x: x.lower() != 'false', default=True)
    parser.add_argument("--save_masks", type=lambda x: x.lower() != 'false', default=True)
    parser.add_argument("--extract_fpoc", type=lambda x: x.lower() != 'false', default=True)
    args = parser.parse_args()
    os.makedirs(args.output_dir, exist_ok=True)
    process_videos(**vars(args))


if __name__ == "__main__":
    main()