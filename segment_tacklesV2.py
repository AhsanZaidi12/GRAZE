#!/usr/bin/env python3
"""
Enhanced Tackle Segmentation -SEGMENT_TacklesV2
=====================================================================

ORIGINAL FIXES (v2):
1. Horizontal player detection: Only reject horizontal DUMMIES, not players
2. Random module conflict: Renamed parameter to avoid shadowing
3. Performance optimization: Balanced validation (fewer frames but smarter)
4. Lower validation thresholds (0.25 vs 0.35)
5. Relaxed matching criteria
6. Excel corruption prevention with file locking and CSV backup
7. Storage optimization - selective frame saving
8. Parallel processing support

BUG FIXES (v3 - this version):
FIX 1 — Threshold relaxation was dead code:
    find_grounding_frame_progressive() returned immediately on first grounding
    success. If SAM2 found no FPOC, lower thresholds (0.30, 0.28) and fallback
    prompt sets were never tried. Fixed by collecting ALL candidates across all
    thresholds/positions/prompts, then letting segment_video() iterate them until
    SAM2 finds a valid FPOC.

FIX 2 — OR-condition confidence gate discarded valid FPOC detections:
    The block `elif overall_confidence < 0.35: needs_manual_review = True` OR
    `elif directional < 0.30: needs_manual_review = True` flagged videos as
    needing review even when SAM2 found FPOC correctly, because grounding
    confidence != contact quality. Fixed to AND-condition with lower thresholds
    (< 0.25 AND < 0.20), only flagging truly degenerate detections.

FIX 3 — Wrong diagnostic label on validation rejection:
    `if len(p.validation_detections) < 3` was logging "low consistency (score)"
    instead of the actual check. Fixed to "insufficient matches (N<3)".
"""

import sys
import os
import cv2
import torch
import numpy as np
from pathlib import Path
import argparse
from typing import List, Tuple, Dict, Optional
import json
from datetime import datetime
import random as random_module
from dataclasses import dataclass, field
import traceback
import shutil
import gc
from openpyxl import Workbook, load_workbook
import csv
import fcntl
import time
import tempfile

from sam2.build_sam import build_sam2_video_predictor
from groundingdino.util.inference import load_model, load_image, predict


# ─────────────────────────────────────────────────────────────────────────────
# Data classes
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class DetectedObject:
    box: np.ndarray
    score: float
    label: str
    kind: str
    center: Tuple[float, float]
    area: float
    aspect_ratio: float
    frame_idx: int = 0
    proximity_to_dummy: float = float('inf')
    selection_score: float = 0.0

    def to_dict(self):
        return {
            'box': self.box.tolist() if isinstance(self.box, np.ndarray) else self.box,
            'score': float(self.score),
            'label': self.label,
            'kind': self.kind,
            'center': (float(self.center[0]), float(self.center[1])),
            'area': float(self.area),
            'aspect_ratio': float(self.aspect_ratio),
            'frame_idx': int(self.frame_idx),
            'proximity_to_dummy': float(self.proximity_to_dummy),
            'selection_score': float(self.selection_score),
        }


@dataclass
class ValidatedDetection:
    primary_frame: int
    primary_detection: DetectedObject
    validation_detections: List[DetectedObject] = field(default_factory=list)
    consistency_score: float = 0.0
    validated: bool = False
    player_motion_score: float = 0.0
    directional_motion_score: float = 0.0

    def to_dict(self):
        return {
            'primary_frame': int(self.primary_frame),
            'primary_detection': self.primary_detection.to_dict(),
            'validation_detections': [d.to_dict() for d in self.validation_detections],
            'consistency_score': float(self.consistency_score),
            'validated': bool(self.validated),
            'player_motion_score': float(self.player_motion_score),
            'directional_motion_score': float(self.directional_motion_score)
        }


class NumpyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.integer):  return int(obj)
        if isinstance(obj, np.floating): return float(obj)
        if isinstance(obj, np.ndarray):  return obj.tolist()
        if isinstance(obj, np.bool_):    return bool(obj)
        return super(NumpyEncoder, self).default(obj)


# ─────────────────────────────────────────────────────────────────────────────
# Excel tracker
# ─────────────────────────────────────────────────────────────────────────────

class ExcelTracker:
    """Robust Excel tracker with file locking and CSV backup."""

    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.csv_path   = excel_path.replace('.xlsx', '.csv')
        self.wb = self.sheet1 = self.sheet2 = None
        self._load_or_create()

    def _acquire_lock(self, file_path: str, timeout: int = 30):
        lock_file  = file_path + '.lock'
        start_time = time.time()
        while True:
            try:
                lock_fd = os.open(lock_file, os.O_CREAT | os.O_EXCL | os.O_RDWR)
                fcntl.flock(lock_fd, fcntl.LOCK_EX)
                return lock_fd
            except (OSError, IOError):
                if time.time() - start_time > timeout:
                    raise TimeoutError(f"Could not acquire lock on {file_path}")
                time.sleep(0.5)

    def _release_lock(self, lock_fd: int, lock_file: str):
        try:
            fcntl.flock(lock_fd, fcntl.LOCK_UN)
            os.close(lock_fd)
            if os.path.exists(lock_file):
                os.remove(lock_file)
        except Exception as e:
            print(f"Warning: Error releasing lock: {e}")

    def _load_or_create(self):
        lock_fd   = None
        lock_file = self.excel_path + '.lock'
        try:
            lock_fd = self._acquire_lock(self.excel_path)
            if os.path.exists(self.excel_path):
                try:
                    self.wb = load_workbook(self.excel_path)
                except Exception as e:
                    print(f"Warning: Could not load Excel, creating new: {e}")
                    self.wb = Workbook()
                self.sheet1 = (self.wb["Correctly Segmented"]
                               if "Correctly Segmented" in self.wb.sheetnames
                               else self.wb.create_sheet("Correctly Segmented"))
                self.sheet2 = (self.wb["Need Manual Review"]
                               if "Need Manual Review" in self.wb.sheetnames
                               else self.wb.create_sheet("Need Manual Review"))
            else:
                self.wb = Workbook()
                if "Sheet" in self.wb.sheetnames:
                    del self.wb["Sheet"]
                self.sheet1 = self.wb.create_sheet("Correctly Segmented")
                self.sheet2 = self.wb.create_sheet("Need Manual Review")
            self._ensure_headers()
        finally:
            if lock_fd is not None:
                self._release_lock(lock_fd, lock_file)

    def _ensure_headers(self):
        if self.sheet1.cell(1, 1).value != "Video Filename":
            self.sheet1.insert_rows(1)
            for col, val in enumerate(
                ["Video Filename", "Threshold", "Confidence",
                 "FFBO Frame", "FPOC Frame", "LFBO Frame",
                 "Total Event Frames", "Grounding Candidate"], 1
            ):
                self.sheet1.cell(1, col).value = val
        if self.sheet2.cell(1, 1).value != "Video Filename":
            self.sheet2.insert_rows(1)
            for col, val in enumerate(
                ["Video Filename", "FFBO Frame", "FPOC Frame", "LFBO Frame"], 1
            ):
                self.sheet2.cell(1, col).value = val

    def _find_row_by_video(self, sheet, video_name: str) -> Optional[int]:
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, 1).value == video_name:
                return row
        return None

    def add_correctly_segmented(self, video_name: str, threshold: float,
                                confidence: float, ffbo_frame: int,
                                fpoc_frame: int, lfbo_frame: int,
                                total_frames: int, grounding_candidate: int = 1):
        row = self._find_row_by_video(self.sheet2, video_name)
        if row:
            self.sheet2.delete_rows(row)
        row = self._find_row_by_video(self.sheet1, video_name)
        vals = [video_name, threshold, confidence, ffbo_frame,
                fpoc_frame, lfbo_frame, total_frames, grounding_candidate]
        if row:
            for c, v in enumerate(vals, 1):
                self.sheet1.cell(row, c).value = v
        else:
            self.sheet1.append(vals)

    def add_needs_review(self, video_name: str, ffbo_frame=None,
                         fpoc_frame=None, lfbo_frame=None):
        if self._find_row_by_video(self.sheet1, video_name):
            return
        row = self._find_row_by_video(self.sheet2, video_name)
        vals = [ffbo_frame or "N/A", fpoc_frame or "N/A", lfbo_frame or "N/A"]
        if row:
            for c, v in enumerate(vals, 2):
                self.sheet2.cell(row, c).value = v
        else:
            self.sheet2.append([video_name] + vals)

    def save(self):
        lock_fd   = None
        lock_file = self.excel_path + '.lock'
        try:
            lock_fd    = self._acquire_lock(self.excel_path)
            temp_excel = self.excel_path + '.tmp'
            self.wb.save(temp_excel)
            if os.path.exists(self.excel_path):
                shutil.copy2(self.excel_path, self.excel_path + '.backup')
            shutil.move(temp_excel, self.excel_path)
            self._save_csv_backup()
            print(f"✓ Excel file saved: {self.excel_path}")
            print(f"✓ CSV backup saved: {self.csv_path}")
        except Exception as e:
            print(f"Error saving Excel: {e}")
            traceback.print_exc()
        finally:
            if lock_fd is not None:
                self._release_lock(lock_fd, lock_file)

    def _save_csv_backup(self):
        try:
            with open(self.csv_path, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['=== CORRECTLY SEGMENTED ==='])
                for row in self.sheet1.iter_rows(values_only=True):
                    writer.writerow(row)
                writer.writerow([])
                writer.writerow(['=== NEED MANUAL REVIEW ==='])
                for row in self.sheet2.iter_rows(values_only=True):
                    writer.writerow(row)
        except Exception as e:
            print(f"Warning: Could not save CSV backup: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# Validators
# ─────────────────────────────────────────────────────────────────────────────

class ObjectValidator:

    @staticmethod
    def compute_iou(box1: np.ndarray, box2: np.ndarray) -> float:
        x1_1, y1_1, x2_1, y2_1 = [float(v) for v in box1]
        x1_2, y1_2, x2_2, y2_2 = [float(v) for v in box2]
        x1_i = max(x1_1, x1_2); y1_i = max(y1_1, y1_2)
        x2_i = min(x2_1, x2_2); y2_i = min(y2_1, y2_2)
        if x2_i <= x1_i or y2_i <= y1_i: return 0.0
        inter = (x2_i - x1_i) * (y2_i - y1_i)
        a1    = max(0.0, x2_1 - x1_1) * max(0.0, y2_1 - y1_1)
        a2    = max(0.0, x2_2 - x1_2) * max(0.0, y2_2 - y1_2)
        union = a1 + a2 - inter
        return float(inter / union) if union > 1e-9 else 0.0

    @staticmethod
    def center_distance(det1: DetectedObject, det2: DetectedObject) -> float:
        dx = det1.center[0] - det2.center[0]
        dy = det1.center[1] - det2.center[1]
        return float(np.sqrt(dx * dx + dy * dy))

    @staticmethod
    def size_ratio(det1: DetectedObject, det2: DetectedObject) -> float:
        a1 = max(float(det1.area), 1e-9)
        a2 = max(float(det2.area), 1e-9)
        return float(min(a1, a2) / max(a1, a2))

    @staticmethod
    def match_best(primary: DetectedObject, candidates: List[DetectedObject],
                   iou_threshold: float, distance_threshold: float,
                   min_size_ratio: float = 0.45) -> Tuple[Optional[DetectedObject], float]:
        best = None; best_conf = 0.0
        for cand in candidates:
            if cand.kind != primary.kind: continue
            iou    = ObjectValidator.compute_iou(primary.box, cand.box)
            dist   = ObjectValidator.center_distance(primary, cand)
            sratio = ObjectValidator.size_ratio(primary, cand)
            if sratio < min_size_ratio: continue
            if iou < iou_threshold and dist > distance_threshold: continue
            iou_score  = min(iou / max(iou_threshold, 1e-6), 1.0)
            dist_score = max(0.0, 1.0 - dist / (distance_threshold * 1.25))
            conf       = 0.55 * iou_score + 0.30 * dist_score + 0.15 * sratio
            if conf > best_conf:
                best_conf = conf; best = cand
        return best, float(best_conf)

    @staticmethod
    def validate_across_frames(primary: DetectedObject, matched: List[DetectedObject],
                               min_validations: int = 3, iou_threshold: float = 0.30,
                               distance_threshold: float = 100.0) -> Tuple[bool, float]:
        if not matched: return False, 0.0
        confidences = []
        for m in matched:
            iou        = ObjectValidator.compute_iou(primary.box, m.box)
            dist       = ObjectValidator.center_distance(primary, m)
            sratio     = ObjectValidator.size_ratio(primary, m)
            iou_score  = min(iou / max(iou_threshold, 1e-6), 1.0)
            dist_score = max(0.0, 1.0 - dist / (distance_threshold * 1.25))
            confidences.append(0.55 * iou_score + 0.30 * dist_score + 0.15 * sratio)
        mean_conf = float(np.mean(confidences)) if confidences else 0.0
        return len(matched) >= min_validations, mean_conf

    @staticmethod
    def compute_player_motion(primary: DetectedObject, matched: List[DetectedObject]) -> float:
        if not matched: return 0.0
        movements = []
        for m in matched:
            dx = primary.center[0] - m.center[0]
            dy = primary.center[1] - m.center[1]
            movements.append(np.sqrt(dx * dx + dy * dy))
        return float(min(np.mean(movements) / 200.0, 1.0))

    @staticmethod
    def compute_directional_motion(primary: DetectedObject,
                                   matched: List[DetectedObject],
                                   dummy: DetectedObject) -> float:
        if not matched or dummy is None: return 0.0
        past_frames = [m for m in matched if m.frame_idx < primary.frame_idx]
        if len(past_frames) < 1: return 0.5
        start_pos = (np.array(past_frames[0].center, dtype=np.float32)
                     if len(past_frames) == 1
                     else np.mean([m.center for m in past_frames], axis=0).astype(np.float32))
        end_pos      = np.array(primary.center,  dtype=np.float32)
        dummy_center = np.array(dummy.center,    dtype=np.float32)
        motion_vec   = end_pos - start_pos
        motion_norm  = np.linalg.norm(motion_vec)
        if motion_norm < 1e-6: return 0.5
        to_dummy      = dummy_center - end_pos
        to_dummy_norm = np.linalg.norm(to_dummy)
        if to_dummy_norm < 1e-6: return 0.5
        dot = float(np.dot(motion_vec / motion_norm, to_dummy / to_dummy_norm))
        return float((dot + 1.0) / 2.0)

    @staticmethod
    def make_reference_detection(box: np.ndarray, kind: str,
                                 frame_idx: int) -> DetectedObject:
        x1, y1, x2, y2 = [float(v) for v in box]
        center = ((x1 + x2) / 2.0, (y1 + y2) / 2.0)
        area   = max((x2 - x1) * (y2 - y1), 1e-6)
        ar     = (y2 - y1) / (x2 - x1 + 1e-6)
        return DetectedObject(
            box=np.array([x1, y1, x2, y2], dtype=np.float32),
            score=1.0, label="reference", kind=kind,
            center=center, area=float(area), aspect_ratio=float(ar),
            frame_idx=int(frame_idx)
        )


class ContactValidator:

    @staticmethod
    def compute_mask_overlap(mask1: np.ndarray, mask2: np.ndarray) -> int:
        return int(np.logical_and(mask1, mask2).sum())

    @staticmethod
    def summarize_overlap(overlap_per_frame: Dict[int, int],
                          min_overlap_pixels: int = 1,
                          store_overlap_curve: bool = False,
                          max_overlap: int = 0,
                          fpoc_frame: Optional[int] = None
                          ) -> Tuple[Optional[int], Optional[int], List[int], Dict]:
        if not overlap_per_frame:
            return None, None, [], {
                'first_contact_frame': None, 'last_contact_frame': None,
                'num_contact_frames': 0, 'max_overlap_pixels': 0,
                'contact_end_reason': None,
                'overlap_per_frame': {} if store_overlap_curve else None
            }
        frames_sorted  = sorted(overlap_per_frame.keys())
        contact_frames = [f for f in frames_sorted if overlap_per_frame[f] >= min_overlap_pixels]
        if not contact_frames:
            return None, None, [], {
                'first_contact_frame': None, 'last_contact_frame': None,
                'num_contact_frames': 0,
                'max_overlap_pixels': max(overlap_per_frame.values()),
                'contact_end_reason': None,
                'overlap_per_frame': ({int(k): int(v) for k, v in overlap_per_frame.items()}
                                      if store_overlap_curve else None)
            }
        first_contact   = contact_frames[0]
        max_overlap_val = max(overlap_per_frame.values())
        last_contact    = contact_frames[-1]
        contact_end_reason = "natural_end"
        if fpoc_frame is not None and max_overlap_val > 0:
            overlap_threshold  = max_overlap_val * 0.10
            consecutive_low    = 0
            required_consec    = 3
            for frame_idx in range(first_contact, frames_sorted[-1] + 1):
                frames_after_fpoc = frame_idx - fpoc_frame
                current_overlap   = overlap_per_frame.get(frame_idx, 0)
                if current_overlap < overlap_threshold:
                    consecutive_low += 1
                else:
                    consecutive_low  = 0
                if frames_after_fpoc > 30 and consecutive_low >= required_consec:
                    last_contact       = frame_idx - required_consec
                    contact_end_reason = "dynamic_threshold"
                    break
        contact_frames = [f for f in frames_sorted
                          if f <= last_contact and overlap_per_frame[f] >= min_overlap_pixels]
        return first_contact, last_contact, contact_frames, {
            'first_contact_frame': first_contact,
            'last_contact_frame':  last_contact,
            'num_contact_frames':  len(contact_frames),
            'max_overlap_pixels':  int(max_overlap_val),
            'contact_end_reason':  contact_end_reason,
            'overlap_per_frame':   ({int(k): int(v) for k, v in overlap_per_frame.items()}
                                    if store_overlap_curve else None)
        }


# ─────────────────────────────────────────────────────────────────────────────
# Main segmenter
# ─────────────────────────────────────────────────────────────────────────────

class ImprovedSegmenterAMP:

    GROUNDING_STEP      = 5
    MAX_GROUNDING_OFFSET = 50

    PROMPT_SETS = [
        {
            "name":   "geared_player",
            "player": ("Primary foreground helmeted full-body american football player "
                       "sprinting, forward-leaning running posture, arms pumping, one leg extended"),
            "dummy":  ("red upright rectangular standing tackle dummy pad, "
                       "vertical blocking dummy, red standing tackling dummy")
        },
        {
            "name":   "non_geared_player",
            "player": ("Primary foreground athletic full-body person running, sprinting, "
                       "forward motion, practicing football tackle"),
            "dummy":  "red upright rectangular standing tackle dummy pad, vertical blocking dummy"
        },
        {
            "name":   "generic_player",
            "player": ("person full-body running towards red object, "
                       "athlete in motion, player approaching target"),
            "dummy":  "red vertical object, standing red pad, upright red target"
        }
    ]

    def __init__(self, sam2_checkpoint: str, sam2_config: str,
                 grounding_dino_checkpoint: str, grounding_dino_config: str,
                 device: str = "cuda"):
        self.device  = device
        self.use_amp = (device == "cuda")
        print("=" * 60)
        print("TACKLE SEGMENTATION - OPTIMIZED VERSION (v3 FIXED)")
        print("=" * 60)
        print("✓ FIX 1: Relaxed thresholds now actually exercised (all candidates collected)")
        print("✓ FIX 2: needs_manual_review uses AND-condition, not OR")
        print("✓ FIX 3: Correct diagnostic label on validation rejection")
        print("✓ Horizontal player allowed (only reject horizontal dummies)")
        print("✓ Balanced validation (16 frames, smarter selection)")
        print("✓ Excel corruption prevention with file locking + CSV backup")
        print("✓ Storage optimization - selective frame saving")
        print("=" * 60)
        print("\nLoading Grounding DINO model...")
        self.grounding_model = load_model(
            grounding_dino_config, grounding_dino_checkpoint, device=device)
        self.grounding_model.eval()
        print("✓ Grounding DINO loaded")
        print("\nLoading SAM2 video predictor...")
        self.video_predictor = build_sam2_video_predictor(
            sam2_config, sam2_checkpoint, device=device)
        print("✓ SAM2 loaded")
        self.validator         = ObjectValidator()
        self.contact_validator = ContactValidator()
        if torch.cuda.is_available():
            torch.cuda.empty_cache()

    # ── Low-level detection ───────────────────────────────────────────────────

    def detect_objects(self, image_path: str, text_prompt: str,
                       box_threshold: float = 0.30,
                       text_threshold: float = 0.25
                       ) -> Tuple[torch.Tensor, torch.Tensor, List[str], np.ndarray]:
        try:
            image_source, image = load_image(image_path)
            with torch.cuda.amp.autocast(enabled=self.use_amp):
                boxes, logits, phrases = predict(
                    model=self.grounding_model, image=image, caption=text_prompt,
                    box_threshold=box_threshold, text_threshold=text_threshold,
                    device=self.device
                )
            h, w, _ = image_source.shape
            if boxes is None or len(boxes) == 0:
                if torch.cuda.is_available(): torch.cuda.empty_cache()
                return torch.empty((0, 4)), torch.empty((0,)), [], image_source
            scale       = torch.tensor([w, h, w, h], device=boxes.device, dtype=boxes.dtype)
            boxes_xyxy  = boxes * scale
            boxes_xyxy[:, :2] -= boxes_xyxy[:, 2:] / 2
            boxes_xyxy[:, 2:] += boxes_xyxy[:, :2]
            boxes_xyxy[:, 0::2] = torch.clamp(boxes_xyxy[:, 0::2], 0, w - 1)
            boxes_xyxy[:, 1::2] = torch.clamp(boxes_xyxy[:, 1::2], 0, h - 1)
            valid      = ((boxes_xyxy[:, 2] > boxes_xyxy[:, 0] + 1) &
                          (boxes_xyxy[:, 3] > boxes_xyxy[:, 1] + 1))
            boxes_xyxy = boxes_xyxy[valid]
            logits     = logits[valid]
            phrases    = [p for p, v in zip(phrases, valid.tolist()) if v]
            if torch.cuda.is_available(): torch.cuda.empty_cache()
            return boxes_xyxy, logits, phrases, image_source
        except RuntimeError as e:
            if "out of memory" in str(e):
                if torch.cuda.is_available():
                    torch.cuda.empty_cache(); gc.collect()
                raise
            raise
        finally:
            if torch.cuda.is_available(): torch.cuda.empty_cache()

    def parse_detections(self, boxes, logits, phrases,
                         image_shape: Tuple[int, int],
                         frame_idx: int) -> List[DetectedObject]:
        detections: List[DetectedObject] = []
        h, w = image_shape
        for box, score, label in zip(boxes, logits, phrases):
            box_np = box.detach().cpu().numpy().astype(np.float32)
            x1, y1, x2, y2 = [float(v) for v in box_np]
            x1 = float(np.clip(x1, 0, w - 1)); x2 = float(np.clip(x2, 0, w - 1))
            y1 = float(np.clip(y1, 0, h - 1)); y2 = float(np.clip(y2, 0, h - 1))
            if x2 <= x1 + 1 or y2 <= y1 + 1: continue
            center = ((x1 + x2) / 2.0, (y1 + y2) / 2.0)
            area   = (x2 - x1) * (y2 - y1)
            ar     = (y2 - y1) / (x2 - x1 + 1e-6)
            detections.append(DetectedObject(
                box=np.array([x1, y1, x2, y2], dtype=np.float32),
                score=float(score), label=str(label), kind="unknown",
                center=center, area=float(area), aspect_ratio=float(ar),
                frame_idx=int(frame_idx),
            ))
        return detections

    def classify_detections(self, detections: List[DetectedObject],
                            image_shape: Tuple[int, int]
                            ) -> Tuple[List[DetectedObject], List[DetectedObject]]:
        """Only reject horizontal DUMMIES, not players."""
        h, w = image_shape
        player_candidates: List[DetectedObject] = []
        dummy_candidates:  List[DetectedObject] = []
        edge_margin_x  = w * 0.15
        edge_margin_y  = h * 0.10
        min_player_area = (h * w) * 0.01

        for det in detections:
            label_l       = det.label.lower()
            in_edge_region = (det.center[0] < edge_margin_x or
                              det.center[0] > w - edge_margin_x or
                              det.center[1] < edge_margin_y)
            is_dummy_label  = any(w in label_l for w in
                                  ['dummy','pad','bag','tackle','training','blocking'])
            is_player_label = any(w in label_l for w in
                                  ['player','person','athlete','helmet','football',
                                   'running','runner','sprint','sprinter'])
            is_horizontal   = det.aspect_ratio < 0.8
            is_vertical     = det.aspect_ratio > 2.0

            if is_dummy_label or is_vertical:
                if is_horizontal: continue
                det.kind = "dummy"
                dummy_candidates.append(det)
                continue
            if is_player_label and not in_edge_region and det.area > min_player_area:
                det.kind = "player"
                player_candidates.append(det)

        return player_candidates, dummy_candidates

    def select_best_dummy(self, dummy_candidates: List[DetectedObject],
                          image_shape: Tuple[int, int]) -> Optional[DetectedObject]:
        if not dummy_candidates: return None
        h, w        = image_shape
        img_center  = (w / 2.0, h / 2.0)
        vertical    = [d for d in dummy_candidates if d.aspect_ratio > 2.0]
        if not vertical:
            vertical = [d for d in dummy_candidates if d.aspect_ratio > 1.5]
        if not vertical: return None
        for det in vertical:
            dist = float(np.sqrt(((det.center[0] - img_center[0]) / w) ** 2 +
                                 ((det.center[1] - img_center[1]) / h) ** 2))
            det.selection_score = float(
                det.score * 0.4 + (1.0 - dist) * 0.3 +
                min(det.aspect_ratio / 3.0, 1.0) * 0.3
            )
        return max(vertical, key=lambda x: x.selection_score)

    def select_players_by_proximity(self, player_candidates: List[DetectedObject],
                                    dummy: DetectedObject, image_shape: Tuple[int, int],
                                    top_n: int = 3) -> List[DetectedObject]:
        if not player_candidates: return []
        h, w     = image_shape
        max_dist = float(np.sqrt(w * w + h * h))
        for p in player_candidates:
            dist = float(np.sqrt((p.center[0]-dummy.center[0])**2 +
                                 (p.center[1]-dummy.center[1])**2))
            p.proximity_to_dummy = dist
            p.selection_score    = float(
                (1.0 - dist / max_dist) * 0.60 +
                p.score * 0.25 +
                min(p.area / (h * w * 0.2), 1.0) * 0.15
            )
        return sorted(player_candidates, key=lambda x: x.selection_score, reverse=True)[:top_n]

    def _grounding_search_offsets(self) -> List[int]:
        offsets = [0]
        for d in range(self.GROUNDING_STEP, self.MAX_GROUNDING_OFFSET + 1,
                       self.GROUNDING_STEP):
            offsets.append(d); offsets.append(-d)
        return offsets

    # ── Multi-position grounding (single call) ────────────────────────────────

    def find_grounding_frame_multiposition(
        self, frame_paths: List[str],
        player_prompt: str, dummy_prompt: str,
        box_threshold: float = 0.30, text_threshold: float = 0.25,
        position_fraction: float = 0.5
    ) -> Tuple[int, List[ValidatedDetection], Optional[ValidatedDetection]]:
        """Find grounding frame around a given position fraction."""
        num_frames      = len(frame_paths)
        middle          = int(num_frames * position_fraction)
        combined_prompt = f"{player_prompt} . {dummy_prompt}"
        primary_frame: Optional[int]           = None
        best_dummy:    Optional[DetectedObject] = None
        top_players:   List[DetectedObject]    = []

        for off in self._grounding_search_offsets():
            idx = middle + off
            if idx < 0 or idx >= num_frames: continue
            try:
                boxes, logits, phrases, image = self.detect_objects(
                    frame_paths[idx], combined_prompt, box_threshold, text_threshold)
                if len(boxes) == 0: continue
                shape   = (image.shape[0], image.shape[1])
                dets    = self.parse_detections(boxes, logits, phrases, shape, idx)
                players, dummies = self.classify_detections(dets, shape)
                dummy   = self.select_best_dummy(dummies, shape)
                if dummy is None or not players: continue
                top_players = self.select_players_by_proximity(players, dummy, shape, top_n=3)
                if not top_players: continue
                primary_frame = idx; best_dummy = dummy; break
            except Exception:
                if torch.cuda.is_available(): torch.cuda.empty_cache()
                continue

        if primary_frame is None or best_dummy is None or not top_players:
            if torch.cuda.is_available(): torch.cuda.empty_cache()
            return -1, [], None

        validation_offsets = [3,-3, 5,-5, 8,-8, 10,-10, 15,-15,
                               20,-20, 25,-25, 30,-30, 35,-35]
        validation_frames  = [primary_frame + o for o in validation_offsets
                               if 0 <= primary_frame + o < num_frames]
        val_box_thresh  = max(0.20, box_threshold  - 0.10)
        val_text_thresh = max(0.20, text_threshold - 0.05)

        cached_players: Dict[int, List[DetectedObject]] = {}
        cached_dummies: Dict[int, List[DetectedObject]] = {}

        for vf in validation_frames:
            try:
                boxes, logits, phrases, image = self.detect_objects(
                    frame_paths[vf], combined_prompt, val_box_thresh, val_text_thresh)
                if len(boxes) == 0:
                    cached_players[vf] = []; cached_dummies[vf] = []; continue
                shape = (image.shape[0], image.shape[1])
                dets  = self.parse_detections(boxes, logits, phrases, shape, vf)
                p_c, d_c = self.classify_detections(dets, shape)
                cached_players[vf] = p_c; cached_dummies[vf] = d_c
            except Exception:
                cached_players[vf] = []; cached_dummies[vf] = []
                if torch.cuda.is_available(): torch.cuda.empty_cache()

        dummy_matches: List[DetectedObject] = []
        for vf in validation_frames:
            match, _ = self.validator.match_best(
                best_dummy, cached_dummies.get(vf, []),
                iou_threshold=0.20, distance_threshold=200.0, min_size_ratio=0.35)
            if match is not None: dummy_matches.append(match)

        dummy_validated, dummy_consistency = self.validator.validate_across_frames(
            best_dummy, dummy_matches, min_validations=2,
            iou_threshold=0.20, distance_threshold=200.0)
        validated_dummy = ValidatedDetection(
            primary_frame, best_dummy, dummy_matches, dummy_consistency, dummy_validated)

        validated_players: List[ValidatedDetection] = []
        for p0 in top_players:
            matches: List[DetectedObject] = []
            for vf in validation_frames:
                match, _ = self.validator.match_best(
                    p0, cached_players.get(vf, []),
                    iou_threshold=0.20, distance_threshold=200.0, min_size_ratio=0.35)
                if match is not None: matches.append(match)
            p_valid, p_cons = self.validator.validate_across_frames(
                p0, matches, min_validations=3,
                iou_threshold=0.20, distance_threshold=200.0)
            p_motion      = self.validator.compute_player_motion(p0, matches)
            p_directional = self.validator.compute_directional_motion(p0, matches, best_dummy)
            validated_players.append(ValidatedDetection(
                primary_frame, p0, matches, p_cons, p_valid, p_motion, p_directional))

        if torch.cuda.is_available(): torch.cuda.empty_cache()
        return primary_frame, validated_players, validated_dummy

    # ── FIX 1: Collect ALL candidates, don't return on first success ──────────

    def find_grounding_frame_progressive(
        self, frame_paths: List[str]
    ) -> Tuple[int, List[ValidatedDetection], Optional[ValidatedDetection],
               str, float, List]:
        """
        FIX 1: Returns (frame, players, dummy, prompt_name, threshold, all_candidates).
        all_candidates is a list of every valid grounding across all
        thresholds / positions / prompt-sets, in quality order.
        segment_video() iterates them until SAM2 finds a valid FPOC,
        so lower thresholds are actually exercised when needed.
        """
        print("\n" + "=" * 60)
        print("PHASE 1: Progressive Multi-Position Grounding")
        print("=" * 60)

        positions  = [0.2, 0.4, 0.5, 0.6, 0.8, 0.9]
        thresholds = [0.35, 0.30, 0.28]

        # Collect every valid grounding candidate
        all_candidates: List[Tuple[int, List[ValidatedDetection],
                                   ValidatedDetection, str, float]] = []

        for prompt_set in self.PROMPT_SETS:
            print(f"\n🎯 Trying prompt set: {prompt_set['name']}")
            player_prompt = prompt_set['player']
            dummy_prompt  = prompt_set['dummy']

            for threshold in thresholds:
                print(f"\n📊 Threshold: {threshold:.2f}")

                for position in positions:
                    print(f"\n  Trying position {position:.2f} "
                          f"(frame {int(len(frame_paths) * position)})...")
                    frame, players, dummy = self.find_grounding_frame_multiposition(
                        frame_paths, player_prompt, dummy_prompt,
                        threshold, 0.25, position)

                    if frame == -1 or not players or dummy is None:
                        continue

                    if len(dummy.validation_detections) < 2:
                        print(f"    ⚠️ Dummy validation failed "
                              f"(consistency={dummy.consistency_score:.3f})")
                        continue

                    players_kept: List[ValidatedDetection] = []
                    for p in players:
                        reasons = []
                        # FIX 3: Correct label — this is a match-count check
                        if len(p.validation_detections) < 3:
                            reasons.append(
                                f"insufficient matches ({len(p.validation_detections)}<3)")
                        if p.player_motion_score < 0.08:
                            reasons.append(f"low motion ({p.player_motion_score:.3f})")
                        if p.directional_motion_score < 0.30:
                            reasons.append(
                                f"low directional ({p.directional_motion_score:.3f})")
                        if p.consistency_score < 0.20:
                            reasons.append(
                                f"low consistency ({p.consistency_score:.3f})")
                        if not reasons:
                            players_kept.append(p)
                        else:
                            print(f"    ✓ Found both objects at frame {frame}")
                            print(f"    ⚠️ Rejected: {'; '.join(reasons)}")

                    if players_kept:
                        players_kept.sort(
                            key=lambda p: (p.directional_motion_score,
                                           p.player_motion_score,
                                           p.consistency_score),
                            reverse=True
                        )
                        print(f"    ✓ Found both objects at frame {frame}")
                        print(f"✓ CANDIDATE: {prompt_set['name']}, "
                              f"threshold {threshold:.2f}, position {position:.2f}")
                        print(f"   Player motion:      {players_kept[0].player_motion_score:.3f}")
                        print(f"   Directional motion: {players_kept[0].directional_motion_score:.3f}")
                        print(f"   Consistency:        {players_kept[0].consistency_score:.3f}")
                        all_candidates.append(
                            (frame, players_kept, dummy,
                             prompt_set['name'], threshold)
                        )
                        # FIX 1: Do NOT return — keep collecting so segment_video
                        # can fall back to weaker candidates when SAM2 finds no FPOC

        if not all_candidates:
            print("\n❌ All grounding attempts failed!")
            return -1, [], None, "none", 0.35, []

        print(f"\n✓ {len(all_candidates)} grounding candidate(s) found — "
              f"SAM2 will try each until FPOC is detected")
        best = all_candidates[0]
        return best[0], best[1], best[2], best[3], best[4], all_candidates

    # ── Backward refinement ───────────────────────────────────────────────────

    def enhanced_backward_refinement(
        self, frame_paths, initial_frame, player_prompt, dummy_prompt,
        box_threshold, text_threshold=0.25,
        reference_player_box=None, reference_dummy_box=None
    ) -> int:
        print("\n🔍 Enhanced backward refinement...")
        combined_prompt = f"{player_prompt} . {dummy_prompt}"
        true_first = initial_frame
        misses     = 0
        max_backward = 20
        stop = max(0, initial_frame - max_backward)

        for idx in range(initial_frame - 1, stop - 1, -1):
            try:
                boxes, logits, phrases, image = self.detect_objects(
                    frame_paths[idx], combined_prompt, box_threshold, text_threshold)
                if len(boxes) == 0:
                    misses += 1
                    if misses > 1: break
                    continue
                shape   = (image.shape[0], image.shape[1])
                dets    = self.parse_detections(boxes, logits, phrases, shape, idx)
                players, dummies = self.classify_detections(dets, shape)
                if reference_player_box is not None and reference_dummy_box is not None:
                    ref_p = self.validator.make_reference_detection(
                        reference_player_box, "player", idx)
                    ref_d = self.validator.make_reference_detection(
                        reference_dummy_box,  "dummy",  idx)
                    pm, _ = self.validator.match_best(ref_p, players, 0.25, 150.0)
                    dm, _ = self.validator.match_best(ref_d, dummies, 0.25, 150.0)
                    if pm and dm:
                        true_first = idx; misses = 0
                    else:
                        misses += 1
                        if misses > 1: break
                else:
                    if players and dummies:
                        true_first = idx; misses = 0
                    else:
                        misses += 1
                        if misses > 1: break
            except Exception:
                misses += 1
                if misses > 1: break

        recovered = initial_frame - true_first
        if recovered > 0:
            print(f"   Phase 1: Refined to frame {true_first} (recovered {recovered} frames)")

        print(f"   Phase 2: Gap detection...")
        gap_offsets         = [5, 10, 20, 50]
        earliest_gap_frame  = true_first

        for gap in gap_offsets:
            test_frame = max(0, true_first - gap)
            if test_frame < 0: break
            try:
                boxes, logits, phrases, image = self.detect_objects(
                    frame_paths[test_frame], combined_prompt,
                    box_threshold, text_threshold)
                if len(boxes) == 0: continue
                shape   = (image.shape[0], image.shape[1])
                dets    = self.parse_detections(boxes, logits, phrases, shape, test_frame)
                players, dummies = self.classify_detections(dets, shape)
                objects_found = False
                if reference_player_box is not None and reference_dummy_box is not None:
                    ref_p = self.validator.make_reference_detection(
                        reference_player_box, "player", test_frame)
                    ref_d = self.validator.make_reference_detection(
                        reference_dummy_box,  "dummy",  test_frame)
                    pm, _ = self.validator.match_best(ref_p, players, 0.25, 150.0)
                    dm, _ = self.validator.match_best(ref_d, dummies, 0.25, 150.0)
                    objects_found = (pm is not None and dm is not None)
                else:
                    objects_found = (players and dummies)

                if objects_found:
                    print(f"      ✓ Found both at frame {test_frame} (gap={gap})")
                    earliest_gap_frame = test_frame
                    left = test_frame; right = true_first
                    while right - left > 1:
                        mid = (left + right) // 2
                        boxes, logits, phrases, image = self.detect_objects(
                            frame_paths[mid], combined_prompt,
                            box_threshold, text_threshold)
                        mid_found = False
                        if len(boxes) > 0:
                            shape = (image.shape[0], image.shape[1])
                            dets  = self.parse_detections(
                                boxes, logits, phrases, shape, mid)
                            mp, md = self.classify_detections(dets, shape)
                            if reference_player_box is not None:
                                rp = self.validator.make_reference_detection(
                                    reference_player_box, "player", mid)
                                rd = self.validator.make_reference_detection(
                                    reference_dummy_box,  "dummy",  mid)
                                pm2, _ = self.validator.match_best(rp, mp, 0.25, 150.0)
                                dm2, _ = self.validator.match_best(rd, md, 0.25, 150.0)
                                mid_found = (pm2 is not None and dm2 is not None)
                            else:
                                mid_found = (mp and md)
                        if mid_found:
                            earliest_gap_frame = mid; right = mid
                        else:
                            left = mid
                    true_first = earliest_gap_frame
                    break
            except Exception:
                continue

        total_recovered = initial_frame - true_first
        print(f"   ✓ Final refined frame: {true_first} (recovered {total_recovered} frames)")
        if torch.cuda.is_available(): torch.cuda.empty_cache()
        return true_first

    # ── Main segment_video ────────────────────────────────────────────────────

    def segment_video(self, video_path, output_dir, min_overlap_pixels=1,
                      store_overlap_curve=False, save_visualization=True,
                      save_masks=True, extract_fpoc=True) -> dict:
        """
        FIX 1: Iterates ALL grounding candidates until SAM2 finds FPOC.
        FIX 2: needs_manual_review uses AND-condition on low-confidence gate.
        """
        video_name       = Path(video_path).stem
        video_output_dir = os.path.join(output_dir, video_name)
        os.makedirs(video_output_dir, exist_ok=True)
        print(f"\n{'='*60}\nProcessing: {video_name}\n{'='*60}")

        with tempfile.TemporaryDirectory() as temp_frames_dir:
            print("\nExtracting frames to temporary directory...")
            frame_paths = self._extract_frames(video_path, temp_frames_dir)
            if not frame_paths:
                return {"success": False, "error": "No frames extracted",
                        "video_name": video_name}
            print(f"Extracted {len(frame_paths)} frames")

            try:
                # FIX 1: Unpack 6-tuple; all_candidates drives fallback iteration
                (grounding_frame, validated_players, validated_dummy,
                 prompt_name, threshold, all_candidates) = \
                    self.find_grounding_frame_progressive(frame_paths)

                if grounding_frame == -1:
                    print(f"   ✓ Added '{video_name}' to 'Need Manual Review'")
                    return {"success": False,
                            "error": "No objects detected after all attempts",
                            "video_name": video_name,
                            "needs_manual_review": True}

                best_result = None
                best_idx    = 0

                # FIX 1: Outer loop over ALL grounding candidates
                for cand_num, (grounding_frame, validated_players,
                               validated_dummy, prompt_name,
                               threshold) in enumerate(all_candidates):

                    print(f"\n--- SAM2 attempt {cand_num+1}/{len(all_candidates)}: "
                          f"grounding frame {grounding_frame}, "
                          f"prompt={prompt_name}, threshold={threshold:.2f} ---")

                    prompt_set    = next((p for p in self.PROMPT_SETS
                                          if p['name'] == prompt_name),
                                         self.PROMPT_SETS[0])
                    player_prompt = prompt_set['player']
                    dummy_prompt  = prompt_set['dummy']
                    top_player_box = validated_players[0].primary_detection.box
                    dummy_box      = validated_dummy.primary_detection.box

                    refined_first_frame = self.enhanced_backward_refinement(
                        frame_paths, grounding_frame, player_prompt, dummy_prompt,
                        threshold, 0.25,
                        reference_player_box=top_player_box,
                        reference_dummy_box=dummy_box)

                    players_sorted        = validated_players
                    actual_grounding_frame = grounding_frame

                    if refined_first_frame < grounding_frame:
                        try:
                            boxes, logits, phrases, image = self.detect_objects(
                                frame_paths[refined_first_frame],
                                f"{player_prompt} . {dummy_prompt}", threshold, 0.25)
                            if len(boxes) > 0:
                                shape = (image.shape[0], image.shape[1])
                                dets  = self.parse_detections(
                                    boxes, logits, phrases, shape, refined_first_frame)
                                p_ref, d_ref = self.classify_detections(dets, shape)
                                if p_ref:
                                    orig = players_sorted[0].primary_detection
                                    top_player_box = min(
                                        p_ref,
                                        key=lambda p: np.hypot(
                                            p.center[0]-orig.center[0],
                                            p.center[1]-orig.center[1])
                                    ).box
                                if d_ref:
                                    orig_d = validated_dummy.primary_detection
                                    dummy_box = min(
                                        d_ref,
                                        key=lambda d: np.hypot(
                                            d.center[0]-orig_d.center[0],
                                            d.center[1]-orig_d.center[1])
                                    ).box
                                actual_grounding_frame = refined_first_frame
                        except Exception:
                            pass

                    # Inner loop over player candidates for this grounding
                    for idx, vp in enumerate(players_sorted):
                        seed_frame      = (actual_grounding_frame if idx == 0
                                           else grounding_frame)
                        player_seed_box = (top_player_box if idx == 0
                                           else vp.primary_detection.box)
                        dummy_seed_box  = (dummy_box if idx == 0
                                           else validated_dummy.primary_detection.box)

                        stats = self._propagate_and_analyze(
                            temp_frames_dir, len(frame_paths), seed_frame,
                            player_seed_box, dummy_seed_box,
                            False, None, min_overlap_pixels, store_overlap_curve)

                        overall_confidence = (
                            vp.consistency_score        * 0.3 +
                            vp.player_motion_score      * 0.3 +
                            vp.directional_motion_score * 0.4
                        )

                        result = {
                            "video_name":               video_name,
                            "num_frames":               len(frame_paths),
                            "grounding_frame":          grounding_frame,
                            "refined_first_frame":      refined_first_frame,
                            "actual_grounding_frame":   seed_frame,
                            "grounding_candidate":      cand_num + 1,
                            "player_candidate_index":   idx,
                            "prompt_set_used":          prompt_name,
                            "threshold_used":           float(threshold),
                            "player_validated":         vp.to_dict(),
                            "dummy_validated":          validated_dummy.to_dict(),
                            "overall_confidence":       float(overall_confidence),
                            "first_frame_both_objects": stats.get("first_frame_both_objects"),
                            "last_frame_both_objects":  stats.get("last_frame_both_objects"),
                            "first_contact_frame":      stats.get("first_contact_frame"),
                            "last_contact_frame":       stats.get("last_contact_frame"),
                            "contact_frames":           stats.get("contact_frames", []),
                            "contact_metrics":          stats.get("contact_metrics", {}),
                            "success":                  True,
                            "has_contact":              stats.get("first_contact_frame") is not None,
                            "needs_manual_review":      False,
                            "amp_mode":                 True,
                        }

                        # FIX 2: AND-condition, not OR — only flag truly degenerate
                        # detections for manual review. A low confidence score does
                        # NOT override a successful SAM2 FPOC detection.
                        if not result["has_contact"]:
                            result["needs_manual_review"] = True
                        else:
                            result["needs_manual_review"] = (
                                overall_confidence < 0.25 and
                                vp.directional_motion_score < 0.20
                            )

                        if result["has_contact"] and not result["needs_manual_review"]:
                            best_result = result
                            best_idx    = idx
                            print(f"✓ FPOC found on grounding candidate {cand_num+1}, "
                                  f"player {idx} — stopping search")
                            break  # inner player loop

                        if best_result is None:
                            best_result = result
                            best_idx    = idx

                        if torch.cuda.is_available(): torch.cuda.empty_cache()
                        gc.collect()

                    # FIX 1: If good FPOC found, also break the outer candidate loop
                    if (best_result is not None and
                            best_result["has_contact"] and
                            not best_result["needs_manual_review"]):
                        break

                    if torch.cuda.is_available(): torch.cuda.empty_cache()
                    gc.collect()

                if best_result is None:
                    print(f"   ✓ Added '{video_name}' to 'Need Manual Review'")
                    return {"success": False, "error": "All candidates failed",
                            "video_name": video_name, "needs_manual_review": True}

                # ── Save masks / visualization / FPOC frame ───────────────────
                if save_masks or save_visualization:
                    cand_num_best = best_result["grounding_candidate"] - 1
                    gf, vp_list, vd, pn, thr = all_candidates[cand_num_best]
                    seed    = best_result["actual_grounding_frame"]
                    pb      = (vp_list[best_idx].primary_detection.box
                               if best_idx < len(vp_list)
                               else vp_list[0].primary_detection.box)
                    db      = vd.primary_detection.box
                    ffbo    = best_result.get("first_frame_both_objects")
                    fpoc    = best_result.get("first_contact_frame")
                    lfbo    = best_result.get("last_frame_both_objects")

                    if save_masks and ffbo is not None and lfbo is not None:
                        masks_dir = os.path.join(video_output_dir, "masks")
                        if os.path.exists(masks_dir):
                            shutil.rmtree(masks_dir, ignore_errors=True)
                        os.makedirs(masks_dir, exist_ok=True)
                        self._propagate_and_analyze(
                            temp_frames_dir, len(frame_paths), seed,
                            pb, db, True, masks_dir,
                            min_overlap_pixels, False,
                            save_range=(ffbo, lfbo))
                        best_result["masks_dir"] = masks_dir

                    if (save_visualization and
                            ffbo is not None and fpoc is not None and lfbo is not None):
                        vis_dir = os.path.join(video_output_dir, "visualization")
                        if os.path.exists(vis_dir):
                            shutil.rmtree(vis_dir, ignore_errors=True)
                        os.makedirs(vis_dir, exist_ok=True)
                        if save_masks:
                            masks_dir = os.path.join(video_output_dir, "masks")
                            self._save_key_frame_visualizations(
                                frame_paths, masks_dir, vis_dir,
                                [ffbo, fpoc, lfbo], ffbo, fpoc, lfbo,
                                ["Player", "Dummy"])
                        best_result["visualization_dir"] = vis_dir
                        print(f"✓ Visualizations saved: "
                              f"FFBO({ffbo}), FPOC({fpoc}), LFBO({lfbo})")

                if extract_fpoc and best_result.get("first_contact_frame") is not None:
                    fpoc_dir = os.path.join(output_dir, "fpoc_frames")
                    os.makedirs(fpoc_dir, exist_ok=True)
                    fi  = best_result["first_contact_frame"]
                    src = os.path.join(temp_frames_dir, f"{fi:05d}.jpg")
                    dst = os.path.join(fpoc_dir,
                                       f"{video_name}_fpoc_frame{fi:05d}.jpg")
                    if os.path.exists(src):
                        shutil.copy2(src, dst)
                        best_result["fpoc_frame_path"] = dst
                        print(f"✓ FPOC frame extracted: {dst}")

                with open(os.path.join(video_output_dir, "metadata.json"), 'w') as f:
                    json.dump(best_result, f, indent=2, cls=NumpyEncoder)

                cand_best = all_candidates[best_result["grounding_candidate"] - 1]
                vp_best   = cand_best[1][best_idx] if best_idx < len(cand_best[1]) \
                             else cand_best[1][0]
                print(f"\n✓ SUCCESS")
                print(f"Overall Confidence:  {best_result['overall_confidence']:.3f}")
                print(f"Threshold used:      {best_result['threshold_used']:.2f}")
                print(f"Grounding candidate: {best_result['grounding_candidate']}"
                      f"/{len(all_candidates)}")
                print(f"FFBO:  Frame {best_result.get('first_frame_both_objects', 'N/A')}")
                print(f"FPOC:  Frame {best_result.get('first_contact_frame',      'N/A')}")
                print(f"LFBO:  Frame {best_result.get('last_frame_both_objects',  'N/A')}")
                print(f"Needs Manual Review: {best_result.get('needs_manual_review', False)}")
                if best_result.get('needs_manual_review'):
                    print(f"   ✓ Added '{video_name}' to 'Need Manual Review'")
                else:
                    print(f"   ✓ Added '{video_name}' to 'Correctly Segmented'")
                return best_result

            except Exception as e:
                traceback.print_exc()
                return {"success": False, "error": str(e),
                        "video_name": video_name, "needs_manual_review": True}
            finally:
                if torch.cuda.is_available(): torch.cuda.empty_cache()
                gc.collect()

    # ── SAM2 propagation ─────────────────────────────────────────────────────

    def _propagate_and_analyze(self, frames_dir, num_frames, seed_frame,
                               player_box, dummy_box,
                               save_masks, masks_dir,
                               min_overlap_pixels, store_overlap_curve,
                               save_range: Optional[Tuple[int, int]] = None) -> dict:
        player_box = player_box.astype(np.float32)
        dummy_box  = dummy_box.astype(np.float32)
        inference_state = self.video_predictor.init_state(
            video_path=frames_dir, async_loading_frames=True)
        self.video_predictor.add_new_points_or_box(
            inference_state=inference_state, frame_idx=int(seed_frame),
            obj_id=0, box=player_box)
        self.video_predictor.add_new_points_or_box(
            inference_state=inference_state, frame_idx=int(seed_frame),
            obj_id=1, box=dummy_box)

        if save_masks and masks_dir:
            os.makedirs(os.path.join(masks_dir, "object_0"), exist_ok=True)
            os.makedirs(os.path.join(masks_dir, "object_1"), exist_ok=True)

        first_both = last_both = None
        overlap_per_frame: Dict[int, int] = {}
        max_overlap = 0

        for out_frame_idx, out_obj_ids, out_mask_logits in \
                self.video_predictor.propagate_in_video(inference_state):
            mask_map: Dict[int, np.ndarray] = {}
            for i, out_obj_id in enumerate(out_obj_ids):
                oid = int(out_obj_id)
                if oid not in (0, 1): continue
                mask = (out_mask_logits[i] > 0.0).detach().cpu().numpy() \
                                                   .squeeze().astype(bool)
                mask_map[oid] = mask
                if save_masks and masks_dir:
                    if save_range is None or \
                            (save_range[0] <= int(out_frame_idx) <= save_range[1]):
                        cv2.imwrite(
                            os.path.join(masks_dir, f"object_{oid}",
                                         f"{int(out_frame_idx):05d}.png"),
                            (mask.astype(np.uint8) * 255))

            has_both = (0 in mask_map) and (1 in mask_map)
            if has_both:
                if first_both is None: first_both = int(out_frame_idx)
                last_both = int(out_frame_idx)
                ov = self.contact_validator.compute_mask_overlap(mask_map[0], mask_map[1])
                overlap_per_frame[int(out_frame_idx)] = int(ov)
                max_overlap = max(max_overlap, ov)

            if int(out_frame_idx) % 50 == 0 and torch.cuda.is_available():
                torch.cuda.empty_cache()

        first_contact_temp, _, _, _ = self.contact_validator.summarize_overlap(
            overlap_per_frame, min_overlap_pixels, False, max_overlap, None)
        first_contact, last_contact, contact_frames, contact_metrics = \
            self.contact_validator.summarize_overlap(
                overlap_per_frame, min_overlap_pixels, store_overlap_curve,
                max_overlap, first_contact_temp)

        return {
            "first_frame_both_objects": first_both,
            "last_frame_both_objects":  last_both,
            "first_contact_frame":      first_contact,
            "last_contact_frame":       last_contact,
            "contact_frames":           contact_frames,
            "contact_metrics":          contact_metrics,
        }

    # ── Frame extraction ──────────────────────────────────────────────────────

    def _extract_frames(self, video_path, output_dir) -> List[str]:
        out      = Path(output_dir)
        existing = sorted(out.glob("*.jpg"))
        if len(existing) > 5: return [str(p) for p in existing]
        for p in existing:
            try: p.unlink()
            except: pass
        cap = cv2.VideoCapture(video_path)
        if not cap.isOpened(): return []
        paths: List[str] = []; idx = 0
        while True:
            ret, frame = cap.read()
            if not ret: break
            fp = os.path.join(output_dir, f"{idx:05d}.jpg")
            if cv2.imwrite(fp, frame):
                paths.append(fp); idx += 1
            else: break
        cap.release()
        return paths

    # ── Key-frame visualization ───────────────────────────────────────────────

    def _save_key_frame_visualizations(self, frame_paths, masks_dir, output_dir,
                                       key_frames, ffbo, fpoc, lfbo, object_labels):
        colors    = [(0, 255, 255), (255, 0, 0)]
        obj0_dir  = Path(masks_dir) / "object_0"
        obj1_dir  = Path(masks_dir) / "object_1"
        if not obj0_dir.exists() or not obj1_dir.exists(): return

        for frame_idx in key_frames:
            if frame_idx < 0 or frame_idx >= len(frame_paths): continue
            frame = cv2.imread(frame_paths[frame_idx])
            if frame is None: continue
            m0_path = obj0_dir / f"{frame_idx:05d}.png"
            m1_path = obj1_dir / f"{frame_idx:05d}.png"
            if not m0_path.exists() or not m1_path.exists(): continue
            m0 = cv2.imread(str(m0_path), cv2.IMREAD_GRAYSCALE)
            m1 = cv2.imread(str(m1_path), cv2.IMREAD_GRAYSCALE)
            if m0 is None or m1 is None: continue

            masks   = {0: (m0 > 0).astype(np.uint8),
                       1: (m1 > 0).astype(np.uint8)}
            overlay = frame.copy()
            is_ffbo = (frame_idx == ffbo)
            is_fpoc = (frame_idx == fpoc)
            is_lfbo = (frame_idx == lfbo)

            for obj_id, mask_binary in masks.items():
                color   = colors[obj_id % len(colors)]
                overlay[mask_binary > 0] = color
                contours, _ = cv2.findContours(
                    mask_binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                if contours:
                    c    = max(contours, key=cv2.contourArea)
                    x, y, w, h = cv2.boundingRect(c)
                    base = (object_labels[obj_id]
                            if obj_id < len(object_labels) else f"Object {obj_id}")
                    lbl  = (f"{base} [FFBO]" if is_ffbo else
                            f"{base} [FPOC]" if is_fpoc else
                            f"{base} [LFBO]" if is_lfbo else base)
                    cv2.rectangle(overlay, (x, y), (x+w, y+h), color, 4)
                    ls, _ = cv2.getTextSize(lbl, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)
                    cv2.rectangle(overlay, (x, y-ls[1]-10), (x+ls[0], y), color, -1)
                    cv2.putText(overlay, lbl, (x, y-5),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255,255,255), 2)

            vis = cv2.addWeighted(frame, 0.6, overlay, 0.4, 0)
            title = ("*** FIRST FRAME WITH BOTH OBJECTS (FFBO) ***" if is_ffbo else
                     "*** FIRST POINT OF CONTACT (FPOC) ***"        if is_fpoc else
                     "*** LAST FRAME WITH BOTH OBJECTS (LFBO) ***")
            color = (255, 0, 255) if is_lfbo else (0, 255, 255)
            cv2.putText(vis, title, (10, 30),
                        cv2.FONT_HERSHEY_SIMPLEX, 1.0, color, 3)
            cv2.imwrite(os.path.join(output_dir, f"{frame_idx:05d}.jpg"), vis)


# ─────────────────────────────────────────────────────────────────────────────
# Process loop
# ─────────────────────────────────────────────────────────────────────────────

def process_videos(video_dir, output_dir, num_test=None, random_selection=False,
                   specific_video=None, process_all=False,
                   batch_mode=False, batch_id=0, batch_size=100,
                   sam2_checkpoint="weights/sam2.1_hiera_large.pt",
                   grounding_checkpoint="weights/groundingdino_swint_ogc.pth",
                   grounding_config="weights/GroundingDINO_SwinT_OGC.py",
                   min_overlap_pixels=1, store_overlap_curve=False,
                   save_visualization=True, save_masks=True, extract_fpoc=True):

    segmenter = ImprovedSegmenterAMP(
        sam2_checkpoint, "configs/sam2.1/sam2.1_hiera_l.yaml",
        grounding_checkpoint, grounding_config,
        "cuda" if torch.cuda.is_available() else "cpu"
    )

    video_files = sorted([f for f in os.listdir(video_dir)
                          if f.endswith(('.mp4','.avi','.mov', '.mkv', '.mpeg','.mod','.MP4','.AVI','.MOV','.MKV', '.MPEG','.MOD'))])

    excel_filename = (f"video_tracking_batch{batch_id}.xlsx" if batch_mode
                      else "video_tracking.xlsx")
    excel_path = os.path.join(output_dir, excel_filename)
    tracker    = ExcelTracker(excel_path)

    if specific_video is not None:
        selected = [f for f in video_files if specific_video in f]
    elif batch_mode:
        start_idx = batch_id * batch_size
        end_idx   = min(start_idx + batch_size, len(video_files))
        selected  = video_files[start_idx:end_idx]
        print(f"\n{'='*60}")
        print(f"BATCH {batch_id}: Processing videos {start_idx} to {end_idx-1}")
        print(f"Total videos in batch: {len(selected)}")
        print(f"{'='*60}\n")
    elif process_all:
        selected = video_files
    elif random_selection and num_test:
        selected = random_module.sample(video_files, min(num_test, len(video_files)))
    elif num_test:
        selected = video_files[:num_test]
    else:
        selected = video_files[:2]

    results = []
    for i, vf in enumerate(selected):
        print(f"\n{'='*60}\nVIDEO {i+1}/{len(selected)}: {vf}\n{'='*60}")
        res = segmenter.segment_video(
            os.path.join(video_dir, vf), output_dir,
            min_overlap_pixels=min_overlap_pixels,
            store_overlap_curve=store_overlap_curve,
            save_visualization=save_visualization,
            save_masks=save_masks, extract_fpoc=extract_fpoc)
        results.append(res)
        video_name = res.get("video_name", vf)

        if res.get("success") and not res.get("needs_manual_review"):
            ffbo  = res.get("first_frame_both_objects", 0)
            fpoc  = res.get("first_contact_frame", 0)
            lfbo  = res.get("last_frame_both_objects", 0)
            total = (lfbo - ffbo) if (ffbo is not None and lfbo is not None) else 0
            tracker.add_correctly_segmented(
                video_name=video_name,
                threshold=res.get("threshold_used", 0.0),
                confidence=res.get("overall_confidence", 0.0),
                ffbo_frame=ffbo, fpoc_frame=fpoc, lfbo_frame=lfbo,
                total_frames=total,
                grounding_candidate=res.get("grounding_candidate", 1))
        else:
            tracker.add_needs_review(
                video_name=video_name,
                ffbo_frame=res.get("first_frame_both_objects"),
                fpoc_frame=res.get("first_contact_frame"),
                lfbo_frame=res.get("last_frame_both_objects"))
        tracker.save()
        print(f"\n📊 Excel file saved: {excel_path}")
        if torch.cuda.is_available(): torch.cuda.empty_cache()
        gc.collect()

    summary_path = os.path.join(
        output_dir, f"summary_batch{batch_id if batch_mode else 0}.json")
    with open(summary_path, 'w') as f:
        json.dump({
            "timestamp":                datetime.now().isoformat(),
            "batch_id":                 batch_id if batch_mode else 0,
            "num_videos_tested":        len(results),
            "success_rate":             (sum(1 for r in results if r.get('success'))
                                         / len(results)) if results else 0,
            "contact_detection_rate":   (sum(1 for r in results if r.get('has_contact'))
                                         / len(results)) if results else 0,
            "correctly_segmented_rate": (sum(1 for r in results
                                             if r.get('success') and
                                             not r.get('needs_manual_review'))
                                         / len(results)) if results else 0,
            "results": results
        }, f, indent=2, cls=NumpyEncoder)


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Enhanced tackle segmentation - v3 FIXED")
    parser.add_argument("--video_dir",            type=str, required=True)
    parser.add_argument("--output_dir",           type=str, required=True)
    parser.add_argument("--num_test",             type=int, default=None)
    parser.add_argument("--random",               action="store_true",
                        dest="random_selection")
    parser.add_argument("--specific_video",       type=str, default=None)
    parser.add_argument("--process_all",          action="store_true")
    parser.add_argument("--batch_mode",           action="store_true")
    parser.add_argument("--batch_id",             type=int, default=0)
    parser.add_argument("--batch_size",           type=int, default=100)
    parser.add_argument("--sam2_checkpoint",      type=str,
                        default="weights/sam2.1_hiera_large.pt")
    parser.add_argument("--grounding_checkpoint", type=str,
                        default="weights/groundingdino_swint_ogc.pth")
    parser.add_argument("--grounding_config",     type=str,
                        default="weights/GroundingDINO_SwinT_OGC.py")
    parser.add_argument("--min_overlap_pixels",   type=int, default=1)
    parser.add_argument("--store_overlap_curve",  action="store_true")
    parser.add_argument("--save_visualization",
                        type=lambda x: x.lower() != 'false', default=True)
    parser.add_argument("--save_masks",
                        type=lambda x: x.lower() != 'false', default=True)
    parser.add_argument("--extract_fpoc",
                        type=lambda x: x.lower() != 'false', default=True)
    args = parser.parse_args()
    os.makedirs(args.output_dir, exist_ok=True)
    process_videos(**vars(args))


if __name__ == "__main__":
    main()