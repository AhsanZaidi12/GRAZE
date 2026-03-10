#!/usr/bin/env python3
"""
Baseline B3 — Motion Analysis Only (No Temporal Validation, No Refinement)
===========================================================================

FIXES APPLIED IN THIS VERSION
──────────────────────────────
FIX 1 — Neutral fallback was indistinguishable from valid motion (B3 == B1):
    compute_player_motion_score() and compute_directional_motion_score()
    returned 0.5 when no history was available. With all three candidates
    getting 0.5/0.5, the combined score reduced to:
        combined = 0.5*0.5 + 0.5*0.3 + prox*0.2 = 0.40 + prox*0.2
    Winner was always the closest player — identical to B1 proximity selection.
    Fixed: return PENALTY (0.10) when no history; return PARTIAL_PENALTY (0.20)
    when fewer than 2 detections are found (insufficient for reliable estimate).
    This makes motion evidence required rather than optional.

FIX 2 — History collection silently collapsed for early primary frames:
    HISTORY_OFFSETS [5,10,15,20,25,30] all clamped to max(0, primary-off).
    For a primary frame of e.g. 10, offsets of 15/20/25/30 all became frame 0,
    and the set() call deduplicated them to {0,5,10}. With only 2-3 unique
    frames and relaxed thresholds, history often came back empty.
    Fixed:
      (a) Only generate offsets where primary_frame - off >= MIN_HISTORY_FRAME (3),
          avoiding frame 0 collapse entirely.
      (b) Added forward history fallback: if backward history yields < 2 matches,
          collect forward frames [+5,+10,+15,+20,+25,+30] and NEGATE the motion
          vector before computing directional alignment. Forward positions after
          primary represent "where the player went" — reversing gives "where they
          came from", which is still directional information.
      (c) Added `motion_fired` flag logged per candidate so you can verify in
          the metadata whether motion actually contributed.

FIX 3 — Proximity normalization used hardcoded 1920×1080 diagonal:
    For portrait-orientation or lower-resolution videos, all proximity scores
    were compressed into a narrow range, making the proximity term
    non-discriminative. Fixed by passing actual image_shape into
    select_players_by_proximity_and_motion() and computing real diagonal.

FIX 4 — Dead variable `h_img` was assigned but never used:
    Removed entirely.

FIX 5 — top_players initial sort used proximity only (no score):
    The initial candidate list fed into motion scoring was sorted by raw distance,
    but proximity_to_dummy was computed correctly. No bug, but clarified with
    explicit selection_score initialization to 0.0 before motion pass.

ABLATION SUMMARY TABLE
┌─────────────────────────────────────┬──────┬──────┬──────┬──────────────┐
│ Component                           │  B1  │  B2  │  B3  │ Full System  │
├─────────────────────────────────────┼──────┼──────┼──────┼──────────────┤
│ Single prompt / fixed threshold     │  ✓   │  ✓   │  ✓   │              │
│ Multi-prompt progressive search     │      │      │      │      ✓       │
│ Multi-frame temporal validation     │      │  ✓   │      │      ✓       │
│ Player motion analysis              │      │      │  ✓   │      ✓       │
│ Backward FFBO refinement            │      │  ✓   │      │      ✓       │
│ Candidate quality sorting + retry   │      │      │      │      ✓       │
└─────────────────────────────────────┴──────┴──────┴──────┴──────────────┘

PURPOSE OF B3
─────────────
B3 isolates the contribution of PLAYER MOTION ANALYSIS — whether knowing the
player's direction of movement and magnitude improves player selection and
ultimately FPOC accuracy.

Added over B1:
  • Backward trajectory history (offsets 5–30 frames before primary co-detection).
    Forward history fallback if backward is sparse (< 2 matches).
  • player_motion_score: mean displacement from history positions to primary,
    normalized by actual video diagonal, clipped to [0,1].
  • directional_motion_score: cosine similarity between player motion vector
    and the vector from player to dummy, mapped to [0,1].
  • Player selection WEIGHTED: directional (0.5) + motion magnitude (0.3)
    + proximity (0.2). Uses actual video resolution for proximity normalization.
  • Penalty scores (0.10) when no history at all; partial penalty (0.20) when
    fewer than 2 history detections found.

NOT added vs full system:
  • No multi-frame validation filter.
  • No backward FFBO refinement.
  • No multi-threshold or multi-prompt fallback.
  • No multi-position search.
  • No candidate quality sorting or multi-SAM2 retry.

WHAT B3 MEASURES
────────────────
  Comparing B3 vs B1 → contribution of motion-aware player selection alone.
  Comparing Full System vs B3 → contribution of temporal validation + refinement.
  Comparing B3 vs B2 → motion vs. temporal consistency (direct component swap).

OUTPUT FILES
────────────
  • <output_dir>/<video_name>/metadata.json   (includes motion_fired flag)
  • <output_dir>/<video_name>/masks/object_0/
  • <output_dir>/<video_name>/masks/object_1/
  • <output_dir>/<video_name>/visualization/
  • <output_dir>/fpoc_frames/
  • <output_dir>/b3_tracking.xlsx + .csv
  • <output_dir>/b3_summary_batch<N>.json
"""

import sys
sys.path.insert(0, '/homes/ahsanzaidi/GroundingDINO')

import os
import cv2
import torch
import numpy as np
from pathlib import Path
import argparse
from typing import List, Tuple, Dict, Optional
import json
from datetime import datetime
import random as random_module
from dataclasses import dataclass, field
import traceback
import shutil
import gc
from openpyxl import Workbook, load_workbook
import csv
import fcntl
import time
import tempfile

from sam2.build_sam import build_sam2_video_predictor
from groundingdino.util.inference import load_model, load_image, predict


# ─────────────────────────────────────────────────────────────────────────────
# Data classes (identical to full system)
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class DetectedObject:
    box: np.ndarray
    score: float
    label: str
    kind: str
    center: Tuple[float, float]
    area: float
    aspect_ratio: float
    frame_idx: int = 0
    proximity_to_dummy: float = float('inf')
    selection_score: float = 0.0

    def to_dict(self):
        return {
            'box': self.box.tolist() if isinstance(self.box, np.ndarray) else self.box,
            'score': float(self.score),
            'label': self.label,
            'kind': self.kind,
            'center': (float(self.center[0]), float(self.center[1])),
            'area': float(self.area),
            'aspect_ratio': float(self.aspect_ratio),
            'frame_idx': int(self.frame_idx),
            'proximity_to_dummy': float(self.proximity_to_dummy),
            'selection_score': float(self.selection_score),
        }


@dataclass
class ValidatedDetection:
    primary_frame: int
    primary_detection: DetectedObject
    validation_detections: List[DetectedObject] = field(default_factory=list)
    consistency_score: float = 0.0
    validated: bool = False
    player_motion_score: float = 0.0
    directional_motion_score: float = 0.0

    def to_dict(self):
        return {
            'primary_frame': int(self.primary_frame),
            'primary_detection': self.primary_detection.to_dict(),
            'validation_detections': [d.to_dict() for d in self.validation_detections],
            'consistency_score': float(self.consistency_score),
            'validated': bool(self.validated),
            'player_motion_score': float(self.player_motion_score),
            'directional_motion_score': float(self.directional_motion_score)
        }


class NumpyEncoder(json.JSONEncoder):
    def default(self, obj):
        if isinstance(obj, np.integer):   return int(obj)
        if isinstance(obj, np.floating):  return float(obj)
        if isinstance(obj, np.ndarray):   return obj.tolist()
        if isinstance(obj, np.bool_):     return bool(obj)
        return super(NumpyEncoder, self).default(obj)


# ─────────────────────────────────────────────────────────────────────────────
# Excel tracker (identical to full system)
# ─────────────────────────────────────────────────────────────────────────────

class ExcelTracker:
    """Robust Excel tracker with file locking and CSV backup."""

    CORRECT_HEADERS = [
        "Video Filename", "Threshold", "Confidence",
        "FFBO Frame", "FPOC Frame", "LFBO Frame",
        "Total Event Frames", "Grounding Candidate"
    ]
    REVIEW_HEADERS = ["Video Filename", "FFBO Frame", "FPOC Frame", "LFBO Frame"]

    def __init__(self, excel_path: str):
        self.excel_path = excel_path
        self.csv_path   = excel_path.replace('.xlsx', '.csv')
        self.wb = self.sheet1 = self.sheet2 = None
        self._load_or_create()

    def _acquire_lock(self, file_path: str, timeout: int = 30):
        lock_file  = file_path + '.lock'
        start_time = time.time()
        while True:
            try:
                lock_fd = os.open(lock_file, os.O_CREAT | os.O_EXCL | os.O_RDWR)
                fcntl.flock(lock_fd, fcntl.LOCK_EX)
                return lock_fd
            except (OSError, IOError):
                if time.time() - start_time > timeout:
                    raise TimeoutError(f"Could not acquire lock on {file_path}")
                time.sleep(0.5)

    def _release_lock(self, lock_fd: int, lock_file: str):
        try:
            fcntl.flock(lock_fd, fcntl.LOCK_UN)
            os.close(lock_fd)
            if os.path.exists(lock_file): os.remove(lock_file)
        except Exception as e:
            print(f"Warning: Error releasing lock: {e}")

    def _load_or_create(self):
        lock_fd = None; lock_file = self.excel_path + '.lock'
        try:
            lock_fd = self._acquire_lock(self.excel_path)
            if os.path.exists(self.excel_path):
                try:   self.wb = load_workbook(self.excel_path)
                except Exception as e:
                    print(f"Warning: Could not load Excel, creating new: {e}")
                    self.wb = Workbook()
                self.sheet1 = (self.wb["Correctly Segmented"]
                               if "Correctly Segmented" in self.wb.sheetnames
                               else self.wb.create_sheet("Correctly Segmented"))
                self.sheet2 = (self.wb["Need Manual Review"]
                               if "Need Manual Review" in self.wb.sheetnames
                               else self.wb.create_sheet("Need Manual Review"))
            else:
                self.wb = Workbook()
                if "Sheet" in self.wb.sheetnames: del self.wb["Sheet"]
                self.sheet1 = self.wb.create_sheet("Correctly Segmented")
                self.sheet2 = self.wb.create_sheet("Need Manual Review")
            self._ensure_headers()
        finally:
            if lock_fd is not None: self._release_lock(lock_fd, lock_file)

    def _ensure_headers(self):
        if self.sheet1.cell(1, 1).value != "Video Filename":
            self.sheet1.insert_rows(1)
            for col, val in enumerate(self.CORRECT_HEADERS, 1):
                self.sheet1.cell(1, col).value = val
        if self.sheet2.cell(1, 1).value != "Video Filename":
            self.sheet2.insert_rows(1)
            for col, val in enumerate(self.REVIEW_HEADERS, 1):
                self.sheet2.cell(1, col).value = val

    def _find_row_by_video(self, sheet, video_name: str) -> Optional[int]:
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row, 1).value == video_name: return row
        return None

    def add_correctly_segmented(self, video_name, threshold, confidence,
                                ffbo_frame, fpoc_frame, lfbo_frame,
                                total_frames, grounding_candidate=1):
        row = self._find_row_by_video(self.sheet2, video_name)
        if row: self.sheet2.delete_rows(row)
        row  = self._find_row_by_video(self.sheet1, video_name)
        vals = [video_name, threshold, confidence, ffbo_frame,
                fpoc_frame, lfbo_frame, total_frames, grounding_candidate]
        if row:
            for c, v in enumerate(vals, 1): self.sheet1.cell(row, c).value = v
        else:
            self.sheet1.append(vals)

    def add_needs_review(self, video_name, ffbo_frame=None,
                         fpoc_frame=None, lfbo_frame=None):
        if self._find_row_by_video(self.sheet1, video_name): return
        row  = self._find_row_by_video(self.sheet2, video_name)
        vals = [ffbo_frame or "N/A", fpoc_frame or "N/A", lfbo_frame or "N/A"]
        if row:
            for c, v in enumerate(vals, 2): self.sheet2.cell(row, c).value = v
        else:
            self.sheet2.append([video_name] + vals)

    def save(self):
        lock_fd = None; lock_file = self.excel_path + '.lock'
        try:
            lock_fd    = self._acquire_lock(self.excel_path)
            temp_excel = self.excel_path + '.tmp'
            self.wb.save(temp_excel)
            if os.path.exists(self.excel_path):
                shutil.copy2(self.excel_path, self.excel_path + '.backup')
            shutil.move(temp_excel, self.excel_path)
            self._save_csv_backup()
            print(f"✓ Excel saved: {self.excel_path}")
        except Exception as e:
            print(f"Error saving Excel: {e}"); traceback.print_exc()
        finally:
            if lock_fd is not None: self._release_lock(lock_fd, lock_file)

    def _save_csv_backup(self):
        try:
            with open(self.csv_path, 'w', newline='') as f:
                w = csv.writer(f)
                w.writerow(['=== CORRECTLY SEGMENTED ==='])
                for row in self.sheet1.iter_rows(values_only=True): w.writerow(row)
                w.writerow([])
                w.writerow(['=== NEED MANUAL REVIEW ==='])
                for row in self.sheet2.iter_rows(values_only=True): w.writerow(row)
        except Exception as e:
            print(f"Warning: Could not save CSV backup: {e}")


# ─────────────────────────────────────────────────────────────────────────────
# Validators (identical to full system)
# ─────────────────────────────────────────────────────────────────────────────

class ObjectValidator:

    @staticmethod
    def compute_iou(box1: np.ndarray, box2: np.ndarray) -> float:
        x1_1, y1_1, x2_1, y2_1 = [float(v) for v in box1]
        x1_2, y1_2, x2_2, y2_2 = [float(v) for v in box2]
        x1_i = max(x1_1, x1_2); y1_i = max(y1_1, y1_2)
        x2_i = min(x2_1, x2_2); y2_i = min(y2_1, y2_2)
        if x2_i <= x1_i or y2_i <= y1_i: return 0.0
        inter = (x2_i - x1_i) * (y2_i - y1_i)
        a1 = max(0.0, x2_1-x1_1) * max(0.0, y2_1-y1_1)
        a2 = max(0.0, x2_2-x1_2) * max(0.0, y2_2-y1_2)
        union = a1 + a2 - inter
        return float(inter / union) if union > 1e-9 else 0.0

    @staticmethod
    def center_distance(det1: DetectedObject, det2: DetectedObject) -> float:
        dx = det1.center[0] - det2.center[0]
        dy = det1.center[1] - det2.center[1]
        return float(np.sqrt(dx * dx + dy * dy))

    @staticmethod
    def size_ratio(det1: DetectedObject, det2: DetectedObject) -> float:
        a1 = max(float(det1.area), 1e-9); a2 = max(float(det2.area), 1e-9)
        return float(min(a1, a2) / max(a1, a2))

    @staticmethod
    def match_best(primary: DetectedObject, candidates: List[DetectedObject],
                   iou_threshold: float, distance_threshold: float,
                   min_size_ratio: float = 0.45) -> Tuple[Optional[DetectedObject], float]:
        best = None; best_conf = 0.0
        for cand in candidates:
            if cand.kind != primary.kind: continue
            iou    = ObjectValidator.compute_iou(primary.box, cand.box)
            dist   = ObjectValidator.center_distance(primary, cand)
            sratio = ObjectValidator.size_ratio(primary, cand)
            if sratio < min_size_ratio: continue
            if iou < iou_threshold and dist > distance_threshold: continue
            iou_score  = min(iou / max(iou_threshold, 1e-6), 1.0)
            dist_score = max(0.0, 1.0 - dist / (distance_threshold * 1.25))
            conf       = 0.55 * iou_score + 0.30 * dist_score + 0.15 * sratio
            if conf > best_conf: best_conf = conf; best = cand
        return best, float(best_conf)

    @staticmethod
    def make_reference_detection(box: np.ndarray, kind: str, frame_idx: int) -> DetectedObject:
        x1, y1, x2, y2 = [float(v) for v in box]
        center = ((x1+x2)/2.0, (y1+y2)/2.0)
        area   = max((x2-x1)*(y2-y1), 1e-6)
        ar     = (y2-y1)/(x2-x1+1e-6)
        return DetectedObject(
            box=np.array([x1, y1, x2, y2], dtype=np.float32),
            score=1.0, label="reference", kind=kind,
            center=center, area=float(area), aspect_ratio=float(ar),
            frame_idx=int(frame_idx))


class ContactValidator:

    @staticmethod
    def compute_mask_overlap(mask1: np.ndarray, mask2: np.ndarray) -> int:
        return int(np.logical_and(mask1, mask2).sum())

    @staticmethod
    def summarize_overlap(overlap_per_frame: Dict[int, int],
                          min_overlap_pixels: int = 1,
                          store_overlap_curve: bool = False,
                          max_overlap: int = 0,
                          fpoc_frame: Optional[int] = None
                          ) -> Tuple[Optional[int], Optional[int], List[int], Dict]:
        if not overlap_per_frame:
            return None, None, [], {
                'first_contact_frame': None, 'last_contact_frame': None,
                'num_contact_frames': 0, 'max_overlap_pixels': 0,
                'contact_end_reason': None,
                'overlap_per_frame': {} if store_overlap_curve else None}

        frames_sorted  = sorted(overlap_per_frame.keys())
        contact_frames = [f for f in frames_sorted if overlap_per_frame[f] >= min_overlap_pixels]

        if not contact_frames:
            return None, None, [], {
                'first_contact_frame': None, 'last_contact_frame': None,
                'num_contact_frames': 0,
                'max_overlap_pixels': max(overlap_per_frame.values()),
                'contact_end_reason': None,
                'overlap_per_frame': ({int(k): int(v) for k, v in overlap_per_frame.items()}
                                      if store_overlap_curve else None)}

        first_contact      = contact_frames[0]
        max_overlap_val    = max(overlap_per_frame.values())
        last_contact       = contact_frames[-1]
        contact_end_reason = "natural_end"

        if fpoc_frame is not None and max_overlap_val > 0:
            overlap_threshold = max_overlap_val * 0.10
            consecutive_low   = 0
            for frame_idx in range(first_contact, frames_sorted[-1] + 1):
                frames_after_fpoc = frame_idx - fpoc_frame
                current_overlap   = overlap_per_frame.get(frame_idx, 0)
                if current_overlap < overlap_threshold: consecutive_low += 1
                else:                                   consecutive_low  = 0
                if frames_after_fpoc > 30 and consecutive_low >= 3:
                    last_contact       = frame_idx - 3
                    contact_end_reason = "dynamic_threshold"
                    break

        contact_frames = [f for f in frames_sorted
                          if f <= last_contact and overlap_per_frame[f] >= min_overlap_pixels]
        return first_contact, last_contact, contact_frames, {
            'first_contact_frame': first_contact,
            'last_contact_frame':  last_contact,
            'num_contact_frames':  len(contact_frames),
            'max_overlap_pixels':  int(max_overlap_val),
            'contact_end_reason':  contact_end_reason,
            'overlap_per_frame':   ({int(k): int(v) for k, v in overlap_per_frame.items()}
                                    if store_overlap_curve else None)}


# ─────────────────────────────────────────────────────────────────────────────
# B3 Segmenter — Motion Analysis Only (all bugs fixed)
# ─────────────────────────────────────────────────────────────────────────────

class B3SegmenterAMP:
    """
    B3 (FIXED): Motion analysis only (no temporal validation, no backward refinement).

    Pipeline:
      1. Sequential frame scan (step=5) to find primary co-detection.
      2. Single fixed prompt ("geared_player") and threshold (0.35).
      3. Backward trajectory history: detect at offsets [-5,-10,-15,-20,-25,-30]
         frames before primary. Only generates offsets where frame >= MIN_HISTORY_FRAME
         to avoid degenerate collapse to frame 0.
      4. Forward history fallback: if backward history yields < MIN_HISTORY_MATCHES,
         collect forward frames [+5,+10,...,+30] and negate the motion vector before
         computing directional alignment.
      5. Motion scores per player candidate (FIXED penalties):
           player_motion_score:      mean displacement / video_diagonal, [0,1].
                                     Returns PENALTY_NO_HISTORY (0.10) if no history.
                                     Returns PENALTY_FEW_HISTORY (0.20) if < 2 matches.
           directional_motion_score: cosine similarity of motion vector → dummy,
                                     mapped to [0,1]. Same penalties as above.
      6. Player selection WEIGHTED:
           directional_motion (0.5) + motion_magnitude (0.3) + proximity (0.2).
           Proximity uses ACTUAL video diagonal (not hardcoded 1920×1080).
      7. SAM2 propagates from the primary co-detection frame (no refinement).
    """

    GROUNDING_STEP      = 5
    FIXED_THRESHOLD     = 0.35
    TEXT_THRESHOLD      = 0.25
    # Relaxed thresholds for history collection only
    HISTORY_BOX_THRESH  = 0.25
    HISTORY_TEXT_THRESH = 0.20
    # FIX 2: Min frame index to generate a history offset for
    # (avoids collapse of multiple offsets to frame 0)
    MIN_HISTORY_FRAME   = 3
    # Offsets (in frames) before / after primary for history collection
    BACKWARD_OFFSETS    = [5, 10, 15, 20, 25, 30]
    FORWARD_OFFSETS     = [5, 10, 15, 20, 25, 30]
    # FIX 2: Minimum matches to use motion scores; below this use partial penalty
    MIN_HISTORY_MATCHES = 2
    # FIX 1: Penalty scores (replaces neutral 0.5 default)
    PENALTY_NO_HISTORY  = 0.10   # no history frames detected at all
    PENALTY_FEW_HISTORY = 0.20   # history found but < MIN_HISTORY_MATCHES

    PLAYER_PROMPT = (
        "Primary foreground helmeted full-body american football player "
        "sprinting, forward-leaning running posture, arms pumping, one leg extended")
    DUMMY_PROMPT  = (
        "red upright rectangular standing tackle dummy pad, "
        "vertical blocking dummy, red standing tackling dummy")

    PLAYER_OBJ_ID = 0
    DUMMY_OBJ_ID  = 1
    OBJECT_LABELS = {0: "Player", 1: "Dummy"}

    def __init__(self, sam2_checkpoint: str, sam2_config: str,
                 grounding_dino_checkpoint: str, grounding_dino_config: str,
                 device: str = "cuda"):
        self.device  = device
        self.use_amp = (device == "cuda")
        print("=" * 60)
        print("BASELINE B3 (FIXED) — Motion Analysis Only")
        print("=" * 60)
        print("✗ Multi-frame temporal validation  → DISABLED")
        print("✓ Player motion analysis           → ENABLED (bugs fixed)")
        print("  FIX 1: Penalty defaults (0.10/0.20) replace neutral 0.5")
        print("  FIX 2: No-history collapse prevented; forward fallback added")
        print("  FIX 3: Proximity uses actual video diagonal, not 1920×1080")
        print("✗ Backward FFBO refinement         → DISABLED")
        print("✗ Multi-threshold / multi-prompt   → DISABLED")
        print("✓ Single prompt + fixed threshold  → ENABLED")
        print("✓ Motion-weighted player selection → ENABLED")
        print("   directional (0.5) + magnitude (0.3) + proximity (0.2)")
        print(f"✓ Object IDs: {self.PLAYER_OBJ_ID}=Player, {self.DUMMY_OBJ_ID}=Dummy")
        print("=" * 60)

        print("\nLoading Grounding DINO model...")
        self.grounding_model = load_model(
            grounding_dino_config, grounding_dino_checkpoint, device=device)
        self.grounding_model.eval()
        print("✓ Grounding DINO loaded")

        print("\nLoading SAM2 video predictor...")
        self.video_predictor = build_sam2_video_predictor(
            sam2_config, sam2_checkpoint, device=device)
        print("✓ SAM2 loaded")

        self.validator         = ObjectValidator()
        self.contact_validator = ContactValidator()
        if torch.cuda.is_available(): torch.cuda.empty_cache()

    # ── Low-level detection ───────────────────────────────────────────────────

    def detect_objects(self, image_path: str, text_prompt: str,
                       box_threshold: float, text_threshold: float
                       ) -> Tuple[torch.Tensor, torch.Tensor, List[str], np.ndarray]:
        try:
            image_source, image = load_image(image_path)
            with torch.cuda.amp.autocast(enabled=self.use_amp):
                boxes, logits, phrases = predict(
                    model=self.grounding_model, image=image, caption=text_prompt,
                    box_threshold=box_threshold, text_threshold=text_threshold,
                    device=self.device)
            h, w, _ = image_source.shape
            if boxes is None or len(boxes) == 0:
                if torch.cuda.is_available(): torch.cuda.empty_cache()
                return torch.empty((0, 4)), torch.empty((0,)), [], image_source
            scale      = torch.tensor([w, h, w, h], device=boxes.device, dtype=boxes.dtype)
            boxes_xyxy = boxes * scale
            boxes_xyxy[:, :2] -= boxes_xyxy[:, 2:] / 2
            boxes_xyxy[:, 2:] += boxes_xyxy[:, :2]
            boxes_xyxy[:, 0::2] = torch.clamp(boxes_xyxy[:, 0::2], 0, w - 1)
            boxes_xyxy[:, 1::2] = torch.clamp(boxes_xyxy[:, 1::2], 0, h - 1)
            valid      = ((boxes_xyxy[:, 2] > boxes_xyxy[:, 0] + 1) &
                          (boxes_xyxy[:, 3] > boxes_xyxy[:, 1] + 1))
            boxes_xyxy = boxes_xyxy[valid]; logits = logits[valid]
            phrases    = [p for p, v in zip(phrases, valid.tolist()) if v]
            if torch.cuda.is_available(): torch.cuda.empty_cache()
            return boxes_xyxy, logits, phrases, image_source
        except RuntimeError as e:
            if "out of memory" in str(e):
                if torch.cuda.is_available(): torch.cuda.empty_cache(); gc.collect()
                raise
            raise
        finally:
            if torch.cuda.is_available(): torch.cuda.empty_cache()

    def parse_detections(self, boxes, logits, phrases,
                         image_shape: Tuple[int, int],
                         frame_idx: int) -> List[DetectedObject]:
        detections: List[DetectedObject] = []
        h, w = image_shape
        for box, score, label in zip(boxes, logits, phrases):
            box_np = box.detach().cpu().numpy().astype(np.float32)
            x1, y1, x2, y2 = [float(v) for v in box_np]
            x1 = float(np.clip(x1, 0, w-1)); x2 = float(np.clip(x2, 0, w-1))
            y1 = float(np.clip(y1, 0, h-1)); y2 = float(np.clip(y2, 0, h-1))
            if x2 <= x1 + 1 or y2 <= y1 + 1: continue
            center = ((x1+x2)/2.0, (y1+y2)/2.0)
            area   = (x2-x1)*(y2-y1)
            ar     = (y2-y1)/(x2-x1+1e-6)
            detections.append(DetectedObject(
                box=np.array([x1, y1, x2, y2], dtype=np.float32),
                score=float(score), label=str(label), kind="unknown",
                center=center, area=float(area), aspect_ratio=float(ar),
                frame_idx=int(frame_idx)))
        return detections

    def classify_detections(self, detections: List[DetectedObject],
                            image_shape: Tuple[int, int]
                            ) -> Tuple[List[DetectedObject], List[DetectedObject]]:
        h, w = image_shape
        player_candidates: List[DetectedObject] = []
        dummy_candidates:  List[DetectedObject] = []
        edge_margin_x   = w * 0.15
        edge_margin_y   = h * 0.10
        min_player_area = (h * w) * 0.01

        for det in detections:
            label_l = det.label.lower()
            in_edge = (det.center[0] < edge_margin_x or
                       det.center[0] > w - edge_margin_x or
                       det.center[1] < edge_margin_y)
            is_dummy_label  = any(kw in label_l for kw in
                                  ['dummy', 'pad', 'bag', 'tackle', 'training', 'blocking'])
            is_player_label = any(kw in label_l for kw in
                                  ['player', 'person', 'athlete', 'helmet', 'football',
                                   'running', 'runner', 'sprint', 'sprinter'])
            is_horizontal = det.aspect_ratio < 0.8
            is_vertical   = det.aspect_ratio > 2.0

            if is_dummy_label or is_vertical:
                if is_horizontal: continue
                det.kind = "dummy"; dummy_candidates.append(det); continue
            if is_player_label and not in_edge and det.area > min_player_area:
                det.kind = "player"; player_candidates.append(det)

        return player_candidates, dummy_candidates

    def select_best_dummy(self, dummy_candidates: List[DetectedObject],
                          image_shape: Tuple[int, int]) -> Optional[DetectedObject]:
        if not dummy_candidates: return None
        h, w = image_shape; ic = (w/2.0, h/2.0)
        vert = [d for d in dummy_candidates if d.aspect_ratio > 2.0]
        if not vert: vert = [d for d in dummy_candidates if d.aspect_ratio > 1.5]
        if not vert: return None
        for det in vert:
            dist = float(np.sqrt(((det.center[0]-ic[0])/w)**2 +
                                 ((det.center[1]-ic[1])/h)**2))
            det.selection_score = float(
                det.score * 0.4 + (1.0 - dist) * 0.3 +
                min(det.aspect_ratio / 3.0, 1.0) * 0.3)
        return max(vert, key=lambda x: x.selection_score)

    # ── FIX 1 + FIX 2: Motion scoring with penalty defaults ──────────────────

    def compute_player_motion_score(self, primary: DetectedObject,
                                    history: List[DetectedObject],
                                    video_diagonal: float) -> float:
        """
        Mean displacement from history positions to primary, normalized by
        actual video diagonal.

        FIX 1: Returns PENALTY_NO_HISTORY (0.10) when history is empty.
                Returns PENALTY_FEW_HISTORY (0.20) when < MIN_HISTORY_MATCHES.
                These replace the old neutral 0.5 that collapsed B3 to B1.
        FIX 3: Uses actual video_diagonal instead of hardcoded 1920×1080.
        """
        if not history:
            return self.PENALTY_NO_HISTORY
        if len(history) < self.MIN_HISTORY_MATCHES:
            return self.PENALTY_FEW_HISTORY

        movements = []
        for h_det in history:
            dx = primary.center[0] - h_det.center[0]
            dy = primary.center[1] - h_det.center[1]
            movements.append(np.sqrt(dx*dx + dy*dy))

        ref_distance = max(video_diagonal * 0.15, 1.0)  # 15% of diagonal as reference
        return float(min(float(np.mean(movements)) / ref_distance, 1.0))

    def compute_directional_motion_score(self, primary: DetectedObject,
                                         history: List[DetectedObject],
                                         dummy: DetectedObject,
                                         forward: bool = False) -> float:
        """
        Cosine similarity between the player's motion vector (history→primary)
        and the vector from player to dummy, mapped to [0,1].

        FIX 1: Returns PENALTY_NO_HISTORY (0.10) when history is empty.
                Returns PENALTY_FEW_HISTORY (0.20) when < MIN_HISTORY_MATCHES.
        FIX 2: forward=True negates the motion vector for forward history frames,
                so the directional score still measures approach toward the dummy
                even when history is collected after the primary frame.
        """
        if not history or dummy is None:
            return self.PENALTY_NO_HISTORY
        if len(history) < self.MIN_HISTORY_MATCHES:
            return self.PENALTY_FEW_HISTORY

        past_centers = np.array([h.center for h in history], dtype=np.float32)
        start_pos    = np.mean(past_centers, axis=0)
        end_pos      = np.array(primary.center,  dtype=np.float32)
        dummy_center = np.array(dummy.center,    dtype=np.float32)

        # For forward history: player was at primary, moved to history positions.
        # Negate to recover the "approach" direction (toward dummy, before primary).
        motion_vec = end_pos - start_pos
        if forward:
            motion_vec = -motion_vec

        motion_norm = np.linalg.norm(motion_vec)
        if motion_norm < 1e-6:
            return self.PENALTY_FEW_HISTORY  # no movement detected

        to_dummy      = dummy_center - end_pos
        to_dummy_norm = np.linalg.norm(to_dummy)
        if to_dummy_norm < 1e-6:
            return self.PENALTY_FEW_HISTORY  # player is at dummy already

        dot = float(np.dot(motion_vec / motion_norm, to_dummy / to_dummy_norm))
        return float((dot + 1.0) / 2.0)

    # ── FIX 2: History collection with backward + forward fallback ────────────

    def _collect_history(
        self, frame_paths: List[str],
        primary_frame: int,
        top_players: List[DetectedObject],
        num_frames: int,
        combined_prompt: str
    ) -> Tuple[Dict[int, List[DetectedObject]], bool]:
        """
        Collect trajectory history for motion scoring.

        FIX 2a: Only generates backward offsets where primary_frame - offset
                 >= MIN_HISTORY_FRAME, preventing degenerate collapse to frame 0.

        FIX 2b: If backward history yields < MIN_HISTORY_MATCHES for ALL players,
                 falls back to forward frames [+5,+10,...] and returns
                 forward=True so the caller can negate motion vectors.

        Returns (player_history dict, is_forward_fallback).
        """
        def _fetch_history_frames(frame_indices: List[int]
                                  ) -> Dict[int, List[DetectedObject]]:
            per_player: Dict[int, List[DetectedObject]] = {
                i: [] for i in range(len(top_players))}
            for hf in frame_indices:
                try:
                    boxes, logits, phrases, image = self.detect_objects(
                        frame_paths[hf], combined_prompt,
                        self.HISTORY_BOX_THRESH, self.HISTORY_TEXT_THRESH)
                    if len(boxes) == 0: continue
                    shape = (image.shape[0], image.shape[1])
                    dets  = self.parse_detections(boxes, logits, phrases, shape, hf)
                    h_players, _ = self.classify_detections(dets, shape)
                    if not h_players: continue
                    for i, ref_player in enumerate(top_players):
                        best_match, _ = self.validator.match_best(
                            ref_player, h_players,
                            iou_threshold=0.15, distance_threshold=250.0,
                            min_size_ratio=0.30)
                        if best_match is not None:
                            per_player[i].append(best_match)
                except Exception:
                    if torch.cuda.is_available(): torch.cuda.empty_cache()
                    continue
            return per_player

        # ── Backward history (FIX 2a: skip offsets that would go below MIN) ──
        backward_frames = sorted(set([
            primary_frame - off
            for off in self.BACKWARD_OFFSETS
            if primary_frame - off >= self.MIN_HISTORY_FRAME
        ]))

        if backward_frames:
            bwd_history = _fetch_history_frames(backward_frames)
            # Check if any player got enough backward matches
            max_bwd_matches = max(len(v) for v in bwd_history.values()) if bwd_history else 0
            if max_bwd_matches >= self.MIN_HISTORY_MATCHES:
                print(f"  ✓ Backward history: {len(backward_frames)} frames queried, "
                      f"max {max_bwd_matches} matches across players")
                return bwd_history, False
            else:
                print(f"  ⚠️ Backward history sparse (max {max_bwd_matches} matches) "
                      f"— trying forward fallback")
        else:
            print(f"  ⚠️ No valid backward frames "
                  f"(primary={primary_frame} < min_offset+{self.MIN_HISTORY_FRAME})"
                  f" — trying forward fallback")
            bwd_history = {i: [] for i in range(len(top_players))}

        # ── Forward history fallback (FIX 2b) ────────────────────────────────
        forward_frames = sorted(set([
            primary_frame + off
            for off in self.FORWARD_OFFSETS
            if primary_frame + off < num_frames
        ]))

        if forward_frames:
            fwd_history = _fetch_history_frames(forward_frames)
            max_fwd_matches = max(len(v) for v in fwd_history.values()) if fwd_history else 0
            print(f"  {'✓' if max_fwd_matches >= self.MIN_HISTORY_MATCHES else '⚠️'} "
                  f"Forward history: {len(forward_frames)} frames queried, "
                  f"max {max_fwd_matches} matches (forward=True, vector will be negated)")
            return fwd_history, True

        # Both directions exhausted — return whatever we have from backward
        print("  ⚠️ Both backward and forward history empty — penalty scores will apply")
        return bwd_history, False

    # ── B3 CORE: Sequential scan + motion analysis ────────────────────────────

    def find_grounding_frame_with_motion(
        self, frame_paths: List[str]
    ) -> Tuple[int, Optional[DetectedObject], Optional[DetectedObject],
               float, float, bool, int]:
        """
        B3 grounding: sequential scan + motion-weighted player selection.
        NO validation gating — motion scores used only for player selection.

        Returns:
            primary_frame       : int   — frame where co-detection found
            best_player         : DetectedObject
            best_dummy          : DetectedObject
            player_motion_score : float
            directional_score   : float
            motion_fired        : bool  — True if motion scores used real history
            history_size        : int   — number of history matches for best player
        """
        print("\n" + "=" * 60)
        print("B3 PHASE 1: Sequential Grounding + Motion Analysis (FIXED)")
        print("=" * 60)
        combined_prompt = f"{self.PLAYER_PROMPT} . {self.DUMMY_PROMPT}"
        num_frames      = len(frame_paths)

        # ── Step 1: sequential scan ───────────────────────────────────────────
        primary_frame = None
        best_dummy: Optional[DetectedObject] = None
        top_players:  List[DetectedObject]   = []
        image_shape:  Tuple[int, int]        = (1080, 1920)  # overwritten below

        for idx in range(0, num_frames, self.GROUNDING_STEP):
            try:
                boxes, logits, phrases, image = self.detect_objects(
                    frame_paths[idx], combined_prompt,
                    self.FIXED_THRESHOLD, self.TEXT_THRESHOLD)
                if len(boxes) == 0: continue
                shape   = (image.shape[0], image.shape[1])
                dets    = self.parse_detections(boxes, logits, phrases, shape, idx)
                players, dummies = self.classify_detections(dets, shape)
                dummy   = self.select_best_dummy(dummies, shape)
                if dummy is None or not players: continue
                h, w    = shape
                # Initial sort by proximity only for candidate seeding
                for p in players:
                    dx = p.center[0] - dummy.center[0]
                    dy = p.center[1] - dummy.center[1]
                    p.proximity_to_dummy = float(np.sqrt(dx*dx + dy*dy))
                top_players   = sorted(players, key=lambda p: p.proximity_to_dummy)[:3]
                primary_frame = idx
                best_dummy    = dummy
                image_shape   = shape
                print(f"  Primary co-detection at frame {idx} "
                      f"({len(top_players)} player candidate(s))")
                break
            except Exception:
                if torch.cuda.is_available(): torch.cuda.empty_cache()
                continue

        if primary_frame is None or best_dummy is None:
            print("❌ No primary co-detection found")
            return -1, None, None, 0.0, 0.0, False, 0

        # FIX 3: Compute actual video diagonal from detected frame shape
        h_frame, w_frame = image_shape
        video_diagonal   = float(np.sqrt(w_frame**2 + h_frame**2))

        # ── Step 2: collect history (FIX 2a + 2b) ────────────────────────────
        player_history, is_forward = self._collect_history(
            frame_paths, primary_frame, top_players, num_frames, combined_prompt)

        # ── Step 3: score each candidate ─────────────────────────────────────
        best_player:            Optional[DetectedObject] = None
        best_motion_score:      float = 0.0
        best_directional_score: float = 0.0
        best_combined:          float = -1.0
        best_history_size:      int   = 0
        motion_fired:           bool  = False

        for i, player in enumerate(top_players):
            history = player_history[i]

            # FIX 1 + FIX 3: penalty defaults + actual diagonal
            motion_score      = self.compute_player_motion_score(
                player, history, video_diagonal)
            directional_score = self.compute_directional_motion_score(
                player, history, best_dummy, forward=is_forward)

            # FIX 3: Proximity uses actual video diagonal
            dx        = player.center[0] - best_dummy.center[0]
            dy        = player.center[1] - best_dummy.center[1]
            dist      = float(np.sqrt(dx*dx + dy*dy))
            prox_score = max(0.0, 1.0 - dist / video_diagonal)

            # B3 weighted selection
            combined = (directional_score * 0.5 +
                        motion_score      * 0.3 +
                        prox_score        * 0.2)
            player.selection_score = float(combined)

            this_fired = len(history) >= self.MIN_HISTORY_MATCHES
            print(f"  Player {i}: directional={directional_score:.3f}, "
                  f"motion={motion_score:.3f}, prox={prox_score:.3f}, "
                  f"combined={combined:.3f}, "
                  f"history={len(history)} frames, "
                  f"motion_fired={'YES' if this_fired else 'NO (penalty)'}, "
                  f"direction={'FWD(neg)' if is_forward else 'BWD'}")

            if combined > best_combined:
                best_combined         = combined
                best_player           = player
                best_motion_score     = motion_score
                best_directional_score = directional_score
                best_history_size     = len(history)
                motion_fired          = this_fired

        if best_player is None:
            best_player            = top_players[0]
            best_motion_score      = self.PENALTY_NO_HISTORY
            best_directional_score = self.PENALTY_NO_HISTORY
            best_history_size      = 0
            motion_fired           = False

        print(f"  ✓ Selected player: combined={best_combined:.3f}, "
              f"directional={best_directional_score:.3f}, "
              f"motion={best_motion_score:.3f}, "
              f"motion_fired={'YES' if motion_fired else 'NO'}")

        if torch.cuda.is_available(): torch.cuda.empty_cache()
        return (primary_frame, best_player, best_dummy,
                best_motion_score, best_directional_score,
                motion_fired, best_history_size)

    # ── SAM2 propagation (identical to full system) ───────────────────────────

    def _propagate_and_analyze(self, frames_dir, num_frames, seed_frame,
                               player_box, dummy_box,
                               save_masks, masks_dir,
                               min_overlap_pixels, store_overlap_curve,
                               save_range: Optional[Tuple[int, int]] = None) -> dict:
        player_box = player_box.astype(np.float32)
        dummy_box  = dummy_box.astype(np.float32)

        inference_state = self.video_predictor.init_state(
            video_path=frames_dir, async_loading_frames=True)
        self.video_predictor.add_new_points_or_box(
            inference_state=inference_state, frame_idx=int(seed_frame),
            obj_id=self.PLAYER_OBJ_ID, box=player_box)
        self.video_predictor.add_new_points_or_box(
            inference_state=inference_state, frame_idx=int(seed_frame),
            obj_id=self.DUMMY_OBJ_ID, box=dummy_box)

        if save_masks and masks_dir:
            os.makedirs(os.path.join(masks_dir, "object_0"), exist_ok=True)
            os.makedirs(os.path.join(masks_dir, "object_1"), exist_ok=True)

        first_both = last_both = None
        overlap_per_frame: Dict[int, int] = {}
        max_overlap = 0

        for out_frame_idx, out_obj_ids, out_mask_logits in \
                self.video_predictor.propagate_in_video(inference_state):
            mask_map: Dict[int, np.ndarray] = {}
            for i, out_obj_id in enumerate(out_obj_ids):
                oid  = int(out_obj_id)
                if oid not in (self.PLAYER_OBJ_ID, self.DUMMY_OBJ_ID): continue
                mask = (out_mask_logits[i] > 0.0).detach().cpu().numpy().squeeze().astype(bool)
                mask_map[oid] = mask
                if save_masks and masks_dir:
                    should_save = True
                    if save_range is not None:
                        should_save = save_range[0] <= int(out_frame_idx) <= save_range[1]
                    if should_save:
                        cv2.imwrite(
                            os.path.join(masks_dir, f"object_{oid}",
                                         f"{int(out_frame_idx):05d}.png"),
                            (mask.astype(np.uint8) * 255))

            has_both = (self.PLAYER_OBJ_ID in mask_map) and (self.DUMMY_OBJ_ID in mask_map)
            if has_both:
                if first_both is None: first_both = int(out_frame_idx)
                last_both = int(out_frame_idx)
                ov = self.contact_validator.compute_mask_overlap(
                    mask_map[self.PLAYER_OBJ_ID], mask_map[self.DUMMY_OBJ_ID])
                overlap_per_frame[int(out_frame_idx)] = int(ov)
                max_overlap = max(max_overlap, ov)
            if int(out_frame_idx) % 50 == 0 and torch.cuda.is_available():
                torch.cuda.empty_cache()

        first_contact_temp, _, _, _ = self.contact_validator.summarize_overlap(
            overlap_per_frame, min_overlap_pixels, False, max_overlap, None)
        first_contact, last_contact, contact_frames, contact_metrics = \
            self.contact_validator.summarize_overlap(
                overlap_per_frame, min_overlap_pixels, store_overlap_curve,
                max_overlap, first_contact_temp)
        return {
            "first_frame_both_objects": first_both,
            "last_frame_both_objects":  last_both,
            "first_contact_frame":      first_contact,
            "last_contact_frame":       last_contact,
            "contact_frames":           contact_frames,
            "contact_metrics":          contact_metrics,
        }

    # ── Main segment_video ────────────────────────────────────────────────────

    def segment_video(self, video_path, output_dir, min_overlap_pixels=1,
                      store_overlap_curve=False, save_visualization=True,
                      save_masks=True, extract_fpoc=True) -> dict:
        video_name       = Path(video_path).stem
        video_output_dir = os.path.join(output_dir, video_name)
        os.makedirs(video_output_dir, exist_ok=True)
        print(f"\n{'='*60}\n[B3-FIXED] Processing: {video_name}\n{'='*60}")

        with tempfile.TemporaryDirectory() as temp_frames_dir:
            print("\nExtracting frames...")
            frame_paths = self._extract_frames(video_path, temp_frames_dir)
            if not frame_paths:
                return {"success": False, "error": "No frames extracted",
                        "video_name": video_name}
            print(f"Extracted {len(frame_paths)} frames")

            try:
                (primary_frame, best_player, best_dummy,
                 motion_score, directional_score,
                 motion_fired, history_size) = \
                    self.find_grounding_frame_with_motion(frame_paths)

                if primary_frame == -1:
                    return {"success": False,
                            "error": "No co-detection found",
                            "video_name": video_name,
                            "needs_manual_review": True}

                player_seed_box = best_player.box.astype(np.float32)
                dummy_seed_box  = best_dummy.box.astype(np.float32)

                print(f"\nB3 PHASE 2: SAM2 Propagation from frame {primary_frame}")
                stats = self._propagate_and_analyze(
                    temp_frames_dir, len(frame_paths), primary_frame,
                    player_seed_box, dummy_seed_box,
                    False, None, min_overlap_pixels, store_overlap_curve)

                result = {
                    "video_name":               video_name,
                    "baseline":                 "B3_FIXED",
                    "num_frames":               len(frame_paths),
                    "grounding_frame":          primary_frame,
                    "seed_frame":               primary_frame,
                    "seed_player_box":          player_seed_box.tolist(),
                    "seed_dummy_box":           dummy_seed_box.tolist(),
                    "player_motion_score":      float(motion_score),
                    "directional_motion_score": float(directional_score),
                    # Diagnostic fields to verify motion actually fired
                    "motion_fired":             bool(motion_fired),
                    "history_size":             int(history_size),
                    "threshold_used":           float(self.FIXED_THRESHOLD),
                    "grounding_candidate":      1,
                    "first_frame_both_objects": stats.get("first_frame_both_objects"),
                    "last_frame_both_objects":  stats.get("last_frame_both_objects"),
                    "first_contact_frame":      stats.get("first_contact_frame"),
                    "last_contact_frame":       stats.get("last_contact_frame"),
                    "contact_frames":           stats.get("contact_frames", []),
                    "contact_metrics":          stats.get("contact_metrics", {}),
                    "success":                  True,
                    "has_contact":              stats.get("first_contact_frame") is not None,
                    "needs_manual_review":      stats.get("first_contact_frame") is None,
                    "amp_mode":                 True,
                }

                # ── Save masks + visualization ────────────────────────────────
                if save_masks or save_visualization:
                    ffbo = result.get("first_frame_both_objects")
                    fpoc = result.get("first_contact_frame")
                    lfbo = result.get("last_frame_both_objects")

                    if save_masks and ffbo is not None and lfbo is not None:
                        masks_dir = os.path.join(video_output_dir, "masks")
                        if os.path.exists(masks_dir):
                            shutil.rmtree(masks_dir, ignore_errors=True)
                        os.makedirs(masks_dir, exist_ok=True)
                        self._propagate_and_analyze(
                            temp_frames_dir, len(frame_paths), primary_frame,
                            player_seed_box, dummy_seed_box,
                            True, masks_dir, min_overlap_pixels, False,
                            save_range=(ffbo, lfbo))
                        result["masks_dir"] = masks_dir

                    if (save_visualization and
                            ffbo is not None and fpoc is not None and lfbo is not None):
                        vis_dir = os.path.join(video_output_dir, "visualization")
                        if os.path.exists(vis_dir):
                            shutil.rmtree(vis_dir, ignore_errors=True)
                        os.makedirs(vis_dir, exist_ok=True)
                        if save_masks:
                            masks_dir = os.path.join(video_output_dir, "masks")
                            self._save_key_frame_visualizations(
                                frame_paths, masks_dir, vis_dir,
                                [ffbo, fpoc, lfbo], ffbo, fpoc, lfbo)
                        result["visualization_dir"] = vis_dir
                        print(f"✓ Visualizations: FFBO({ffbo}), FPOC({fpoc}), LFBO({lfbo})")

                if extract_fpoc and result.get("first_contact_frame") is not None:
                    fpoc_dir = os.path.join(output_dir, "fpoc_frames")
                    os.makedirs(fpoc_dir, exist_ok=True)
                    fi  = result["first_contact_frame"]
                    src = os.path.join(temp_frames_dir, f"{fi:05d}.jpg")
                    dst = os.path.join(fpoc_dir, f"{video_name}_fpoc_frame{fi:05d}.jpg")
                    if os.path.exists(src):
                        shutil.copy2(src, dst)
                        result["fpoc_frame_path"] = dst

                with open(os.path.join(video_output_dir, "metadata.json"), 'w') as f:
                    json.dump(result, f, indent=2, cls=NumpyEncoder)

                motion_status = "FIRED" if motion_fired else f"PENALTY (history={history_size})"
                print(f"\n✓ [B3-FIXED] SUCCESS")
                print(f"  FFBO: {result.get('first_frame_both_objects', 'N/A')}")
                print(f"  FPOC: {result.get('first_contact_frame', 'N/A')}")
                print(f"  LFBO: {result.get('last_frame_both_objects', 'N/A')}")
                print(f"  Motion: {motion_status}")
                return result

            except Exception as e:
                traceback.print_exc()
                return {"success": False, "error": str(e),
                        "video_name": video_name, "needs_manual_review": True}
            finally:
                if torch.cuda.is_available(): torch.cuda.empty_cache()
                gc.collect()

    # ── Frame extraction ──────────────────────────────────────────────────────

    def _extract_frames(self, video_path, output_dir) -> List[str]:
        out      = Path(output_dir)
        existing = sorted(out.glob("*.jpg"))
        if len(existing) > 5: return [str(p) for p in existing]
        for p in existing:
            try: p.unlink()
            except: pass
        cap   = cv2.VideoCapture(video_path)
        if not cap.isOpened(): return []
        paths: List[str] = []; idx = 0
        while True:
            ret, frame = cap.read()
            if not ret: break
            fp = os.path.join(output_dir, f"{idx:05d}.jpg")
            if cv2.imwrite(fp, frame): paths.append(fp); idx += 1
            else: break
        cap.release()
        return paths

    # ── Key-frame visualization (identical to full system) ────────────────────

    def _save_key_frame_visualizations(self, frame_paths, masks_dir, output_dir,
                                       key_frames, ffbo, fpoc, lfbo):
        colors   = {self.PLAYER_OBJ_ID: (0, 255, 255),
                    self.DUMMY_OBJ_ID:  (255, 0, 0)}
        obj0_dir = Path(masks_dir) / "object_0"
        obj1_dir = Path(masks_dir) / "object_1"
        if not obj0_dir.exists() or not obj1_dir.exists(): return

        for frame_idx in key_frames:
            if frame_idx < 0 or frame_idx >= len(frame_paths): continue
            frame = cv2.imread(frame_paths[frame_idx])
            if frame is None: continue
            m0_path = obj0_dir / f"{frame_idx:05d}.png"
            m1_path = obj1_dir / f"{frame_idx:05d}.png"
            if not m0_path.exists() or not m1_path.exists(): continue
            m0 = cv2.imread(str(m0_path), cv2.IMREAD_GRAYSCALE)
            m1 = cv2.imread(str(m1_path), cv2.IMREAD_GRAYSCALE)
            if m0 is None or m1 is None: continue

            masks   = {self.PLAYER_OBJ_ID: (m0 > 0).astype(np.uint8),
                       self.DUMMY_OBJ_ID:  (m1 > 0).astype(np.uint8)}
            overlay = frame.copy()
            is_ffbo = (frame_idx == ffbo)
            is_fpoc = (frame_idx == fpoc)
            is_lfbo = (frame_idx == lfbo)

            for obj_id, mask_binary in masks.items():
                color = colors[obj_id]
                overlay[mask_binary > 0] = color
                contours, _ = cv2.findContours(
                    mask_binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
                if contours:
                    c        = max(contours, key=cv2.contourArea)
                    x, y, w, h = cv2.boundingRect(c)
                    base     = self.OBJECT_LABELS[obj_id]
                    lbl      = (f"{base} [FFBO]" if is_ffbo else
                                f"{base} [FPOC]" if is_fpoc else
                                f"{base} [LFBO]" if is_lfbo else base)
                    cv2.rectangle(overlay, (x, y), (x+w, y+h), color, 4)
                    ls, _    = cv2.getTextSize(lbl, cv2.FONT_HERSHEY_SIMPLEX, 0.6, 2)
                    cv2.rectangle(overlay, (x, y-ls[1]-10), (x+ls[0], y), color, -1)
                    cv2.putText(overlay, lbl, (x, y-5),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 2)

            vis   = cv2.addWeighted(frame, 0.6, overlay, 0.4, 0)
            title = ("*** FIRST FRAME WITH BOTH OBJECTS (FFBO) ***" if is_ffbo else
                     "*** FIRST POINT OF CONTACT (FPOC) ***"        if is_fpoc else
                     "*** LAST FRAME WITH BOTH OBJECTS (LFBO) ***")
            color = (255, 0, 255) if is_lfbo else (0, 255, 255)
            if title:
                cv2.putText(vis, title, (10, 30),
                            cv2.FONT_HERSHEY_SIMPLEX, 1.0, color, 3)
            cv2.imwrite(os.path.join(output_dir, f"{frame_idx:05d}.jpg"), vis)


# ─────────────────────────────────────────────────────────────────────────────
# Process loop
# ─────────────────────────────────────────────────────────────────────────────

def process_videos(video_dir, output_dir, num_test=None, random_selection=False,
                   specific_video=None, process_all=False,
                   batch_mode=False, batch_id=0, batch_size=100,
                   sam2_checkpoint="weights/sam2.1_hiera_large.pt",
                   grounding_checkpoint="weights/groundingdino_swint_ogc.pth",
                   grounding_config="weights/GroundingDINO_SwinT_OGC.py",
                   min_overlap_pixels=1, store_overlap_curve=False,
                   save_visualization=True, save_masks=True, extract_fpoc=True):

    segmenter = B3SegmenterAMP(
        sam2_checkpoint, "configs/sam2.1/sam2.1_hiera_l.yaml",
        grounding_checkpoint, grounding_config,
        "cuda" if torch.cuda.is_available() else "cpu")

    video_files = sorted([f for f in os.listdir(video_dir)
                          if f.endswith(('.mp4', '.avi', '.mov', '.mkv', '.mpeg', '.mod',
                                         '.MP4', '.AVI', '.MOV', '.MKV', '.MPEG', '.MOD'))])
    excel_path = os.path.join(output_dir,
                              f"b3_tracking_batch{batch_id}.xlsx" if batch_mode
                              else "b3_tracking.xlsx")
    tracker = ExcelTracker(excel_path)

    if specific_video is not None:
        selected = [f for f in video_files if specific_video in f]
    elif batch_mode:
        start_idx = batch_id * batch_size
        end_idx   = min(start_idx + batch_size, len(video_files))
        selected  = video_files[start_idx:end_idx]
        print(f"\nBATCH {batch_id}: videos {start_idx}–{end_idx-1} ({len(selected)} total)")
    elif process_all:
        selected = video_files
    elif random_selection and num_test:
        selected = random_module.sample(video_files, min(num_test, len(video_files)))
    elif num_test:
        selected = video_files[:num_test]
    else:
        selected = video_files[:2]

    results = []
    for i, vf in enumerate(selected):
        print(f"\n{'='*60}\n[B3-FIXED] VIDEO {i+1}/{len(selected)}: {vf}\n{'='*60}")
        res  = segmenter.segment_video(
            os.path.join(video_dir, vf), output_dir,
            min_overlap_pixels=min_overlap_pixels,
            store_overlap_curve=store_overlap_curve,
            save_visualization=save_visualization,
            save_masks=save_masks, extract_fpoc=extract_fpoc)
        results.append(res)
        vn = res.get("video_name", vf)

        if res.get("success") and not res.get("needs_manual_review"):
            ffbo  = res.get("first_frame_both_objects", 0)
            fpoc  = res.get("first_contact_frame", 0)
            lfbo  = res.get("last_frame_both_objects", 0)
            total = (lfbo - ffbo + 1) if (ffbo is not None and lfbo is not None) else 0
            tracker.add_correctly_segmented(
                vn, res.get("threshold_used", 0.0),
                (res.get("directional_motion_score", 0.0) * 0.5 +
                 res.get("player_motion_score",      0.0) * 0.5),
                ffbo, fpoc, lfbo, total, 1)
        else:
            tracker.add_needs_review(vn,
                res.get("first_frame_both_objects"),
                res.get("first_contact_frame"),
                res.get("last_frame_both_objects"))
        tracker.save()
        if torch.cuda.is_available(): torch.cuda.empty_cache()
        gc.collect()

    # Compute motion_fired rate for summary diagnostics
    motion_fired_rate = (
        sum(1 for r in results if r.get('motion_fired', False)) / len(results)
        if results else 0.0)

    summary_path = os.path.join(output_dir,
                                f"b3_summary_batch{batch_id if batch_mode else 0}.json")
    with open(summary_path, 'w') as f:
        json.dump({
            "timestamp":                datetime.now().isoformat(),
            "baseline":                 "B3_FIXED",
            "description":              (
                "Motion analysis only — no temporal validation, no backward refinement. "
                "Bugs fixed: penalty defaults (0.10/0.20), no-history collapse prevented, "
                "forward fallback, actual video diagonal for proximity normalization."),
            "batch_id":                 batch_id if batch_mode else 0,
            "num_videos_tested":        len(results),
            "success_rate":             (sum(1 for r in results if r.get('success'))
                                         / len(results)) if results else 0,
            "contact_detection_rate":   (sum(1 for r in results if r.get('has_contact'))
                                         / len(results)) if results else 0,
            "correctly_segmented_rate": (sum(1 for r in results
                                             if r.get('success') and
                                             not r.get('needs_manual_review'))
                                         / len(results)) if results else 0,
            # Diagnostic: what fraction of videos actually had motion fire?
            # Should be > 0.5 for a valid B3 ablation. If still ~0, history
            # collection is failing for domain-specific reasons.
            "motion_fired_rate":        float(motion_fired_rate),
            "results":                  results
        }, f, indent=2, cls=NumpyEncoder)
    print(f"\n✓ [B3-FIXED] Summary saved: {summary_path}")
    print(f"  Motion fired rate: {motion_fired_rate:.1%} of videos")
    print(f"  (If < 50%, history collection still failing — investigate primary frame timing)")


def main():
    parser = argparse.ArgumentParser(
        description="Baseline B3 FIXED — Motion Analysis Only (no validation, no refinement)")
    parser.add_argument("--video_dir",            type=str, required=True)
    parser.add_argument("--output_dir",           type=str, required=True)
    parser.add_argument("--num_test",             type=int, default=None)
    parser.add_argument("--random",               action="store_true", dest="random_selection")
    parser.add_argument("--specific_video",       type=str, default=None)
    parser.add_argument("--process_all",          action="store_true")
    parser.add_argument("--batch_mode",           action="store_true")
    parser.add_argument("--batch_id",             type=int, default=0)
    parser.add_argument("--batch_size",           type=int, default=100)
    parser.add_argument("--sam2_checkpoint",      type=str, default="weights/sam2.1_hiera_large.pt")
    parser.add_argument("--grounding_checkpoint", type=str, default="weights/groundingdino_swint_ogc.pth")
    parser.add_argument("--grounding_config",     type=str, default="weights/GroundingDINO_SwinT_OGC.py")
    parser.add_argument("--min_overlap_pixels",   type=int, default=1)
    parser.add_argument("--store_overlap_curve",  action="store_true")
    parser.add_argument("--save_visualization",
                        type=lambda x: x.lower() != 'false', default=True)
    parser.add_argument("--save_masks",
                        type=lambda x: x.lower() != 'false', default=True)
    parser.add_argument("--extract_fpoc",
                        type=lambda x: x.lower() != 'false', default=True)
    args = parser.parse_args()
    os.makedirs(args.output_dir, exist_ok=True)
    process_videos(**vars(args))


if __name__ == "__main__":
    main()