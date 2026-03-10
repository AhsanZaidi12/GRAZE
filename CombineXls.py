"""
combine_results.py
==================
Combines four batch tracking Excel files, joins Ground Truth FPOC,
computes FPOC error, and produces a summary sheet.
"""

import os
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
#  CONFIGURATION  —  change paths here
# ─────────────────────────────────────────────────────────────────────────────
GT_LOOKUP_FILE  = "/ocean/projects/asc180003p/szaidi/videos/gt_fpoc_lookup.xlsx"

# BATCH_DIR       = "/ocean/projects/asc180003p/szaidi/Tackle_Sam2/Results_v3_mp4"  #full result

#Baseline Results path
# BATCH_DIR       = "/ocean/projects/asc180003p/szaidi/Tackle_Sam2/ResultsB1_mp4"
BATCH_DIR       = "/ocean/projects/asc180003p/szaidi/Tackle_Sam2/ResultsB2_mp4"
# BATCH_DIR       = "/ocean/projects/asc180003p/szaidi/Tackle_Sam2/ResultsB3_mp4"

# #full xls
# BATCH_FILES     = [
#     os.path.join(BATCH_DIR, "video_tracking_batch0.xlsx"),
#     os.path.join(BATCH_DIR, "video_tracking_batch1.xlsx"),
#     os.path.join(BATCH_DIR, "video_tracking_batch2.xlsx"),
#     os.path.join(BATCH_DIR, "video_tracking_batch3.xlsx"),
# ]
#Baseline xls , comment for nonbaseline experiments:
BATCH_FILES     = [
    os.path.join(BATCH_DIR, "b2_tracking_batch0.xlsx"),
    os.path.join(BATCH_DIR, "b2_tracking_batch1.xlsx"),
    os.path.join(BATCH_DIR, "b2_tracking_batch2.xlsx"),
    os.path.join(BATCH_DIR, "b2_tracking_batch3.xlsx"),
]



OUTPUT_FILE     = os.path.join(BATCH_DIR, "combined_results.xlsx")

MAX_VIDEO_ID    = 785   # IDs above this are invalid (frame numbers leaking in)
# ─────────────────────────────────────────────────────────────────────────────


# ── Helpers ───────────────────────────────────────────────────────────────────

def norm_id(val):
    """Return zero-padded 3-digit string ID, or None if invalid."""
    if val is None:
        return None
    s = re.sub(r'\.(mp4|avi|mov|mkv|wmv)$', '', str(val).strip(), flags=re.IGNORECASE)
    digits = re.sub(r'\D', '', s)
    if not digits:
        return None
    n = int(digits)
    if n < 1 or n > MAX_VIDEO_ID:
        return None
    return str(n).zfill(3)


def to_int(val):
    """Convert a cell value to int, returning None if not possible."""
    if val is None:
        return None
    try:
        if str(val).strip().upper() in ('N/A', 'NAN', 'NONE', ''):
            return None
        return int(float(str(val)))
    except (ValueError, TypeError):
        return None


# ── Step 1: Load GT FPOC lookup ───────────────────────────────────────────────

def load_gt_lookup(path):
    """
    Reads gt_fpoc_lookup.xlsx.
    Expected columns: #  |  Filename (e.g. 001)  |  Ground Truth FPOC
    Returns dict: '001' → 43
    """
    gt = {}
    if not os.path.exists(path):
        print(f"[ERROR] GT lookup file not found: {path}")
        return gt

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Find Filename and GT FPOC columns by header
    fn_col = gt_col = None
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or '').lower().strip()
        if 'filename' in h or h == '#':
            if fn_col is None and 'filename' in h:
                fn_col = c
        if 'ground' in h and 'fpoc' in h:
            gt_col = c
        if 'fpoc' in h and 'ground' in h:
            gt_col = c

    # Fallback to positional (col B = filename, col C = GT FPOC)
    if fn_col is None: fn_col = 2
    if gt_col is None: gt_col = 3

    for r in range(2, ws.max_row + 1):
        vid = norm_id(ws.cell(r, fn_col).value)
        val = to_int(ws.cell(r, gt_col).value)
        if vid and val is not None:
            gt[vid] = val

    print(f"  GT lookup loaded: {len(gt)} videos with GT FPOC")
    return gt


# ── Step 2: Load and combine batch files ─────────────────────────────────────

def load_batches(batch_files):
    """
    Returns two DataFrames: df_cs (Correctly Segmented), df_nmr (Need Manual Review).
    Skips blank rows and deduplicates by video ID (first occurrence wins).
    """
    CS_COLS  = ['Video Filename', 'Threshold', 'Confidence',
                'FFBO Frame', 'FPOC Frame', 'LFBO Frame', 'Total Event Frames']
    NMR_COLS = ['Video Filename', 'FFBO Frame', 'FPOC Frame', 'LFBO Frame']

    cs_rows, nmr_rows = [], []

    for path in batch_files:
        if not os.path.exists(path):
            print(f"  [WARNING] Not found, skipping: {path}")
            continue

        fname = os.path.basename(path)
        print(f"  Reading {fname} ...")

        try:
            # ── Correctly Segmented ───────────────────────────────────────
            df = pd.read_excel(path, sheet_name='Correctly Segmented',
                               header=0, dtype=str)
            # Remap columns by name (handles extra/missing cols safely)
            col_map = {}
            for std in CS_COLS:
                for actual in df.columns:
                    if std.lower().strip() == str(actual).lower().strip():
                        col_map[std] = actual
                        break
            for std in CS_COLS:
                if std not in col_map:
                    df[std] = None
                    col_map[std] = std

            df = df.rename(columns={v: k for k, v in col_map.items()})[CS_COLS]
            df['_vid'] = df['Video Filename'].apply(norm_id)
            df = df[df['_vid'].notna()].copy()
            cs_rows.append(df)

            # ── Need Manual Review ────────────────────────────────────────
            df2 = pd.read_excel(path, sheet_name='Need Manual Review',
                                header=0, dtype=str)
            col_map2 = {}
            for std in NMR_COLS:
                for actual in df2.columns:
                    if std.lower().strip() == str(actual).lower().strip():
                        col_map2[std] = actual
                        break
            for std in NMR_COLS:
                if std not in col_map2:
                    df2[std] = None
                    col_map2[std] = std

            df2 = df2.rename(columns={v: k for k, v in col_map2.items()})[NMR_COLS]
            df2['_vid'] = df2['Video Filename'].apply(norm_id)
            df2 = df2[df2['_vid'].notna()].copy()
            nmr_rows.append(df2)

        except Exception as e:
            print(f"  [ERROR] {fname}: {e}")

    df_cs  = pd.concat(cs_rows,  ignore_index=True) if cs_rows  else pd.DataFrame(columns=CS_COLS  + ['_vid'])
    df_nmr = pd.concat(nmr_rows, ignore_index=True) if nmr_rows else pd.DataFrame(columns=NMR_COLS + ['_vid'])

    # Deduplicate: keep first occurrence per video ID
    df_cs  = df_cs.drop_duplicates(subset='_vid', keep='first').reset_index(drop=True)
    df_nmr = df_nmr.drop_duplicates(subset='_vid', keep='first').reset_index(drop=True)

    # Videos in CS should not also appear in NMR
    df_nmr = df_nmr[~df_nmr['_vid'].isin(df_cs['_vid'])].reset_index(drop=True)

    print(f"  Combined CS : {len(df_cs)} videos")
    print(f"  Combined NMR: {len(df_nmr)} videos")
    return df_cs, df_nmr


# ── Step 3 & 4: Join GT FPOC and compute FPOC error ──────────────────────────

def add_gt_and_error(df_cs, df_nmr, gt):
    """Adds Ground Truth FPOC and FPOC Error columns to both DataFrames."""

    def join_gt(df, fpoc_col='FPOC Frame'):
        df = df.copy()
        df['Ground Truth FPOC'] = df['_vid'].map(gt)
        fpoc_int = df[fpoc_col].apply(to_int)
        gt_int   = df['Ground Truth FPOC'].apply(to_int)
        df['FPOC Error'] = [
            fpoc - gt_v if (fpoc is not None and gt_v is not None) else None
            for fpoc, gt_v in zip(fpoc_int, gt_int)
        ]
        return df

    df_cs  = join_gt(df_cs,  fpoc_col='FPOC Frame')
    df_nmr = join_gt(df_nmr, fpoc_col='FPOC Frame')
    return df_cs, df_nmr


# ── Step 5: Compute summary stats ────────────────────────────────────────────

def compute_summary(df_cs):
    """Returns a dict of summary statistics for correctly segmented videos."""
    errors = df_cs['FPOC Error'].apply(to_int).dropna().abs()
    total_cs   = len(df_cs)
    total_eval = len(errors)

    def count_pct(threshold):
        n = int((errors < threshold).sum())
        pct = 100.0 * n / total_eval if total_eval > 0 else 0.0
        return n, pct

    return {
        'total_cs':    total_cs,
        'total_eval':  total_eval,
        'lt5':         count_pct(5),
        'lt10':        count_pct(10),
        'lt15':        count_pct(15),
        'lt20':        count_pct(20),
    }


# ── Styling ───────────────────────────────────────────────────────────────────

HDR_FILL   = PatternFill("solid", start_color="1F4E79")
HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
CS_FILL    = PatternFill("solid", start_color="EBF3FB")
NMR_FILL   = PatternFill("solid", start_color="FFF2CC")
ERR_FILL   = PatternFill("solid", start_color="FCE4D6")   # |error| >= 20
OK_FILL    = PatternFill("solid", start_color="E2EFDA")   # |error| < 10
MID_FILL   = PatternFill("solid", start_color="FFEB9C")   # 10 <= |error| < 20
SUM_FILL   = PatternFill("solid", start_color="D9E1F2")

thin       = Side(style='thin', color="BFBFBF")
BORDER     = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER     = Alignment(horizontal='center', vertical='center', wrap_text=False)
LEFT       = Alignment(horizontal='left',   vertical='center')


def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = CENTER; cell.border = BORDER


def style_row(ws, r, ncols, fill):
    for c in range(1, ncols + 1):
        cell = ws.cell(r, c)
        cell.font = Font(name="Arial", size=10)
        cell.fill = fill; cell.border = BORDER
        cell.alignment = CENTER if c != 2 else LEFT


def row_fill_cs(fpoc_error):
    """Choose row background based on absolute FPOC error."""
    if fpoc_error is None: return CS_FILL
    ae = abs(fpoc_error)
    if ae < 10:  return OK_FILL
    if ae < 20:  return MID_FILL
    return ERR_FILL


# ── Write sheet: Correctly Segmented ─────────────────────────────────────────

def write_cs_sheet(wb, df_cs):
    ws = wb.create_sheet("Correctly Segmented")
    headers = ['Video Filename', 'Threshold', 'Confidence',
               'FFBO Frame', 'FPOC Frame', 'LFBO Frame', 'Total Event Frames',
               'Ground Truth FPOC', 'FPOC Error']
    ws.append(headers)
    style_header(ws, len(headers))

    for _, row in df_cs.iterrows():
        fpoc_err = to_int(row.get('FPOC Error'))
        gt_val   = to_int(row.get('Ground Truth FPOC'))
        conf     = row.get('Confidence')
        try:
            conf = round(float(conf), 4)
        except (ValueError, TypeError):
            conf = None

        ws.append([
            row['_vid'],
            row.get('Threshold'),
            conf,
            to_int(row.get('FFBO Frame')),
            to_int(row.get('FPOC Frame')),
            to_int(row.get('LFBO Frame')),
            to_int(row.get('Total Event Frames')),
            gt_val,
            fpoc_err,
        ])
        style_row(ws, ws.max_row, len(headers), row_fill_cs(fpoc_err))

    widths = [16, 12, 13, 13, 13, 13, 18, 18, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ── Write sheet: Need Manual Review ──────────────────────────────────────────

def write_nmr_sheet(wb, df_nmr):
    ws = wb.create_sheet("Need Manual Review")
    headers = ['Video Filename', 'FFBO Frame', 'FPOC Frame', 'LFBO Frame',
               'Ground Truth FPOC', 'FPOC Error']
    ws.append(headers)
    style_header(ws, len(headers))

    for _, row in df_nmr.iterrows():
        fpoc_raw = row.get('FPOC Frame', 'N/A')
        fpoc_int = to_int(fpoc_raw)
        fpoc_err = to_int(row.get('FPOC Error'))
        gt_val   = to_int(row.get('Ground Truth FPOC'))

        ws.append([
            row['_vid'],
            to_int(row.get('FFBO Frame')),
            fpoc_int if fpoc_int is not None else 'N/A',
            to_int(row.get('LFBO Frame')),
            gt_val,
            fpoc_err,
        ])
        fill = ERR_FILL if (fpoc_err is not None and abs(fpoc_err) >= 20) else NMR_FILL
        style_row(ws, ws.max_row, len(headers), fill)

    widths = [16, 13, 13, 13, 18, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


# ── Write sheet: Summary ──────────────────────────────────────────────────────

def write_summary_sheet(wb, stats):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16

    def hdr(title):
        r = ws.max_row + (2 if ws.max_row > 1 else 1)
        ws.cell(r, 1, title).font = Font(name="Arial", bold=True,
                                          size=11, color="1F4E79")
        ws.cell(r, 1).fill      = SUM_FILL
        ws.cell(r, 1).border    = BORDER
        ws.cell(r, 2).fill      = SUM_FILL
        ws.cell(r, 2).border    = BORDER
        ws.cell(r, 3).fill      = SUM_FILL
        ws.cell(r, 3).border    = BORDER
        ws.merge_cells(f"A{r}:C{r}")

    def row(label, count, pct=None):
        r = ws.max_row + 1
        ws.cell(r, 1, label).font      = Font(name="Arial", size=10)
        ws.cell(r, 2, count).font      = Font(name="Arial", size=10, bold=True)
        ws.cell(r, 2).alignment        = CENTER
        ws.cell(r, 1).border           = BORDER
        ws.cell(r, 2).border           = BORDER
        ws.cell(r, 3).border           = BORDER
        if pct is not None:
            ws.cell(r, 3, f"{pct:.1f}%").font      = Font(name="Arial", size=10,
                                                           color="595959")
            ws.cell(r, 3).alignment = CENTER

    # Title
    ws.cell(1, 1, "FPOC Detection Accuracy — Summary Report")
    ws.cell(1, 1).font = Font(name="Arial", bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 24

    hdr("📁  Dataset Overview")
    row("Total correctly segmented videos",   stats['total_cs'])
    row("Videos with evaluable FPOC error",   stats['total_eval'])

    hdr("🎯  FPOC Accuracy (Correctly Segmented)")
    ws.cell(ws.max_row + 1, 3, "% of evaluable").font = Font(name="Arial",
                                                              italic=True, size=9,
                                                              color="808080")
    ws.cell(ws.max_row, 1, "Threshold").font = Font(name="Arial",
                                                     italic=True, size=9,
                                                     color="808080")
    ws.cell(ws.max_row, 2, "Count").font     = Font(name="Arial",
                                                     italic=True, size=9,
                                                     color="808080")

    n5,  p5  = stats['lt5']
    n10, p10 = stats['lt10']
    n15, p15 = stats['lt15']
    n20, p20 = stats['lt20']

    row("|FPOC Error| < 5 frames",   n5,  p5) 
    row("|FPOC Error| < 10 frames",  n10, p10)
    row("|FPOC Error| < 15 frames",  n15, p15)
    row("|FPOC Error| < 20 frames",  n20, p20)
    row("|FPOC Error| ≥ 20 frames (potential wrong detection)",
        stats['total_eval'] - n20,
        100.0 - p20 if stats['total_eval'] > 0 else 0.0)

    hdr("🎨  Row Color Legend")
    for label, fill in [
        ("Green  — |error| < 10 frames",   OK_FILL),
        ("Yellow — 10 ≤ |error| < 20 frames", MID_FILL),
        ("Salmon — |error| ≥ 20 frames",    ERR_FILL),
        ("Blue   — No GT FPOC available",   CS_FILL),
    ]:
        r = ws.max_row + 1
        ws.cell(r, 1, label).font = Font(name="Arial", size=10)
        ws.cell(r, 1).fill        = fill
        ws.cell(r, 1).border      = BORDER
        ws.cell(r, 2).fill        = fill
        ws.cell(r, 2).border      = BORDER
        ws.cell(r, 3).fill        = fill
        ws.cell(r, 3).border      = BORDER
        ws.merge_cells(f"A{r}:C{r}")

    hdr("⚙️  Configuration")
    row("GT FPOC source",   GT_LOOKUP_FILE)
    row("Batch results dir", BATCH_DIR)
    row("Output file",       OUTPUT_FILE)
    row("Max valid video ID", MAX_VIDEO_ID)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Step 1: Loading Ground Truth FPOC ...")
    gt = load_gt_lookup(GT_LOOKUP_FILE)

    print("\nStep 2: Loading and combining batch files ...")
    df_cs, df_nmr = load_batches(BATCH_FILES)

    print("\nStep 3-4: Joining GT FPOC and computing FPOC Error ...")
    df_cs, df_nmr = add_gt_and_error(df_cs, df_nmr, gt)
    cs_with_gt  = df_cs['Ground Truth FPOC'].notna().sum()
    nmr_with_gt = df_nmr['Ground Truth FPOC'].notna().sum()
    print(f"  CS  videos with GT FPOC: {cs_with_gt}/{len(df_cs)}")
    print(f"  NMR videos with GT FPOC: {nmr_with_gt}/{len(df_nmr)}")

    print("\nStep 5: Computing summary statistics ...")
    stats = compute_summary(df_cs)

    print("\nBuilding Excel output ...")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_cs_sheet(wb, df_cs)
    write_nmr_sheet(wb, df_nmr)
    write_summary_sheet(wb, stats)

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)

    n5,  p5  = stats['lt5']
    n10, p10 = stats['lt10']
    n15, p15 = stats['lt15']
    n20, p20 = stats['lt20']
    total_e  = stats['total_eval']

    print(f"\n{'='*60}")
    print(f"✅  Saved: {OUTPUT_FILE}")
    print(f"\n── Summary ───────────────────────────────────────────────")
    print(f"  Correctly Segmented  : {stats['total_cs']}")
    print(f"  Need Manual Review   : {len(df_nmr)}")
    print(f"  Evaluable (CS+GT)    : {total_e}")
    print(f"  |Error| < 5 frames   : {n5:4d}  ({p5:.1f}%)") 
    print(f"  |Error| < 10 frames  : {n10:4d}  ({p10:.1f}%)")
    print(f"  |Error| < 15 frames  : {n15:4d}  ({p15:.1f}%)")
    print(f"  |Error| < 20 frames  : {n20:4d}  ({p20:.1f}%)")
    print(f"  |Error| >= 20 frames : {total_e - n20:4d}  ({100-p20:.1f}%)")


if __name__ == "__main__":
    main()